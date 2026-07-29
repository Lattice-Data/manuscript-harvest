"""Spreadsheets -> table cards. The largest supplement category in the corpus.

172 of the supplementary files here are `.xlsx`, one is a legacy `.xls`, and 18
are `.csv`. Two things this module refuses to do, both learned from the corpus:

- It never calls `calculate_dimension()`. `44161_2025_612_MOESM5_ESM.xlsx` has
  worksheets with no declared dimensions, and openpyxl raises
  `ValueError: Worksheet is unsized` for those; the reader has to reset the
  dimensions and count rows itself.
- It never reads a whole sheet just to find out how big it is. Scanning stops at
  `limits.max_scan_rows` and the card says the profile is partial.
"""

import csv
import io
import warnings
from typing import Any, List, Sequence, Tuple

from . import ooxml, tables
from .limits import Limits

OK = "ok"
NO_TEXT = "no_text"
UNREADABLE = "unreadable"
UNSUPPORTED = "unsupported_format"

_DELIMITERS = ",\t;|"


def _scan_rows(iterator, limits: Limits) -> Tuple[List[Sequence[Any]], bool]:
    """Pull up to the cap, plus one row to learn whether more existed."""
    rows: List[Sequence[Any]] = []
    truncated = False
    for row in iterator:
        if len(rows) >= limits.max_scan_rows:
            truncated = True
            break
        rows.append(tuple(row))
    return rows, truncated


def _silence_openpyxl() -> None:
    """Drop openpyxl's "extension is not supported" warnings for this call.

    Publisher workbooks carry conditional-formatting, data-validation and chart
    extensions that openpyxl warns about and then discards. They concern features
    this stage never reads. The filter has to be installed around the *row
    iteration* as well as the load, because a read-only workbook parses its
    sheets lazily and the warnings fire while scanning.
    """
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


def _load(data: bytes):
    """Open a workbook, with the strict-OOXML retry. Raises on unreadable bytes."""
    import openpyxl

    workbook = openpyxl.load_workbook(
        io.BytesIO(data), read_only=True, data_only=True, keep_links=False)
    if workbook.worksheets:
        return workbook, False
    # Zero worksheets is how openpyxl reports a strict-conformance workbook.
    # Relax the namespaces and try once more before believing it is empty.
    workbook.close()
    relaxed = ooxml.relax_strict(data)
    if relaxed is None:
        return openpyxl.load_workbook(
            io.BytesIO(data), read_only=True, data_only=True, keep_links=False), False
    return openpyxl.load_workbook(
        io.BytesIO(relaxed), read_only=True, data_only=True, keep_links=False), True


def cards_from_xlsx(
    data: bytes, source_file: str, limits: Limits
) -> Tuple[List[tables.TableCard], str, dict]:
    try:
        import openpyxl  # noqa: F401
    except ImportError:  # pragma: no cover - openpyxl is a hard requirement
        return [], UNSUPPORTED, {"reason": "openpyxl is not installed"}

    with warnings.catch_warnings():
        _silence_openpyxl()
        return _cards_from_xlsx(data, source_file, limits)


def _cards_from_xlsx(
    data: bytes, source_file: str, limits: Limits
) -> Tuple[List[tables.TableCard], str, dict]:
    try:
        workbook, relaxed = _load(data)
    except Exception as e:
        return [], UNREADABLE, {"reason": f"{type(e).__name__}: {e}"}

    cards: List[tables.TableCard] = []
    meta: dict = {"sheets": 0, "sheets_skipped": 0}
    if relaxed:
        meta["strict_ooxml"] = True
    try:
        worksheets = list(workbook.worksheets)
        meta["sheets"] = len(worksheets)
        if not worksheets:
            return [], UNREADABLE, {
                **meta,
                "reason": "the workbook declares no worksheets; if it is a strict "
                          "ISO-29500 file the namespace relaxation did not help"}
        if len(worksheets) > limits.max_sheets:
            meta["sheets_skipped"] = len(worksheets) - limits.max_sheets
            worksheets = worksheets[: limits.max_sheets]

        for sheet in worksheets:
            # Read the declared row count first: `reset_dimensions` clears it, and
            # in read-only mode nothing recomputes it, so asking afterwards gives
            # None even for a sheet that said how big it was.
            declared = sheet.max_row
            total = declared if isinstance(declared, int) and declared > 0 else None
            # Unsized worksheets yield nothing until their dimensions are reset.
            try:
                sheet.reset_dimensions()
            except (AttributeError, TypeError, ValueError):
                pass
            try:
                rows, truncated = _scan_rows(sheet.iter_rows(values_only=True), limits)
            except Exception as e:
                meta.setdefault("errors", []).append(
                    f"sheet {sheet.title!r}: {type(e).__name__}: {e}")
                continue

            card = tables.build_card(
                rows,
                source_file=source_file,
                locator=f"sheet {sheet.title!r}",
                limits=limits,
                title=sheet.title,
                n_rows_total=total,
                truncated=truncated,
                data_ref={"file": source_file, "sheet": sheet.title},
            )
            if card is not None:
                cards.append(card)
            if len(cards) >= limits.max_tables_per_file:
                break
    finally:
        workbook.close()

    if not cards:
        return [], NO_TEXT, meta
    return cards, OK, meta


def cards_from_xls(
    data: bytes, source_file: str, limits: Limits
) -> Tuple[List[tables.TableCard], str, dict]:
    """Legacy binary `.xls`. One file in this corpus needs it.

    `xlrd` 2.x reads only this format, which is exactly the reason to keep it
    optional: without it the single file is reported as unsupported rather than
    dragging in a dependency for 1 of 191 spreadsheets.
    """
    try:
        import xlrd
    except ImportError:
        return [], UNSUPPORTED, {
            "reason": "legacy .xls needs the optional xlrd package (pip install xlrd)"}

    try:
        book = xlrd.open_workbook(file_contents=data)
    except Exception as e:
        return [], UNREADABLE, {"reason": f"{type(e).__name__}: {e}"}

    cards: List[tables.TableCard] = []
    meta = {"sheets": book.nsheets}
    for sheet in book.sheets()[: limits.max_sheets]:
        limit = min(sheet.nrows, limits.max_scan_rows)
        rows = [tuple(sheet.row_values(index)) for index in range(limit)]
        card = tables.build_card(
            rows,
            source_file=source_file,
            locator=f"sheet {sheet.name!r}",
            limits=limits,
            title=sheet.name,
            n_rows_total=sheet.nrows,
            truncated=sheet.nrows > limit,
            data_ref={"file": source_file, "sheet": sheet.name},
        )
        if card is not None:
            cards.append(card)
    if not cards:
        return [], NO_TEXT, meta
    return cards, OK, meta


def _sniff_delimiter(sample: str) -> str:
    try:
        return csv.Sniffer().sniff(sample, delimiters=_DELIMITERS).delimiter
    except csv.Error:
        # Fall back to whichever candidate appears most on the first line.
        first = sample.splitlines()[0] if sample.splitlines() else ""
        counts = {d: first.count(d) for d in _DELIMITERS}
        best = max(counts, key=counts.get)
        return best if counts[best] else ","


def cards_from_csv(
    data: bytes, source_file: str, limits: Limits
) -> Tuple[List[tables.TableCard], str, dict]:
    text = data.decode("utf-8-sig", errors="replace")
    if not text.strip():
        return [], NO_TEXT, {}
    delimiter = _sniff_delimiter(text[:8192])
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    try:
        rows, truncated = _scan_rows(reader, limits)
    except csv.Error as e:
        return [], UNREADABLE, {"reason": f"{type(e).__name__}: {e}"}

    card = tables.build_card(
        rows,
        source_file=source_file,
        locator="rows",
        limits=limits,
        n_rows_total=None,
        truncated=truncated,
        data_ref={"file": source_file, "delimiter": delimiter},
    )
    meta = {"delimiter": delimiter}
    if card is None:
        return [], NO_TEXT, meta
    return [card], OK, meta
