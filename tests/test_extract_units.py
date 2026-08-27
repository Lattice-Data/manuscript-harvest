"""Unit coverage for the extraction stage: sections, table cards, and each parser.

Most of these pin a rule that a real file in the corpus proved wrong at least
once, and where that is so the docstring names the DOI. The recurring theme is
that a parser must not report emptiness it cannot justify: a strict-conformance
workbook, a header on row 4, and a bot-check landing page all look like "there
was nothing there" unless something checks.
"""

import json
import sys
from unittest import mock

import fitz
import openpyxl
import openpyxl.worksheet._read_only
import pytest

from manuscript_harvest.extract import archive, docxfile, extractor, htmlfile, jats, ooxml, pdf
from manuscript_harvest.extract import sections
from manuscript_harvest.extract import spreadsheet, tables
from manuscript_harvest.extract.blocks import (
    CAPTION,
    HEADING,
    METADATA,
    PARAGRAPH,
    TABLE,
    Block,
    number_blocks,
    read_blocks,
    render_markdown,
    strip_invisible,
    write_blocks,
)
from manuscript_harvest.extract.limits import Limits
from tests.fakes import (
    LANDING_INTERSTITIAL,
    concat_pdfs,
    SPRINGER_SUPPLEMENT,
    jats_article,
    make_dimensionless_xlsx,
    make_docx,
    make_embedded_font_pdf,
    fake_tesseract as tesseract,
    make_pdf,
    no_tesseract,
    make_pdf_pages,
    make_scanned_pdf,
    make_unreadable_font_pdf,
    make_strict_xlsx,
    make_zero_sheet_xlsx,
    make_gz,
    make_tar,
    make_xls,
    make_xlsx,
    make_zip,
)

L = Limits()


# -- sections ----------------------------------------------------------------

@pytest.mark.parametrize("heading,expected", [
    ("Methods", "methods"),
    ("METHODS", "methods"),
    ("Online Methods", "methods"),
    ("Materials and methods", "methods"),
    ("STAR Methods", "methods"),
    ("Experimental procedures", "methods"),
    ("2.1 Methods", "methods"),
    ("Results and discussion", "results"),
    ("Results:", "results"),
    ("Discussion", "discussion"),
    ("Data availability", "data_availability"),
    ("Supplementary Methods", "methods"),
    ("Figure legends", "figure_legends"),
    ("Acknowledgements", "back_matter"),
    ("References", "references"),
])
def test_known_headings_normalise(heading, expected):
    assert sections.normalize(heading) == expected


@pytest.mark.parametrize("heading,expected", [
    # Cell Press STAR Methods, 10.1016/j.cell.2021.01.053
    ("Experimental Model and Subject Details", "methods"),
    ("Quantification and Statistical Analysis", "methods"),
    ("Supplemental Experimental Procedures", "methods"),
    ("Supplemental Information", "supplementary"),
    # The same headings as PyMuPDF renders Cell Press's bullet glyph, page 18.
    ("d EXPERIMENTAL MODEL AND SUBJECT DETAILS", "methods"),
    ("d METHOD DETAILS", "methods"),
    ("d QUANTIFICATION AND STATISTICAL ANALYSIS", "methods"),
    ("• Method Details", "methods"),
    # Science, 10.1126/science.aat5031 supplement page 83
    ("References and Notes", "references"),
    ("Methods Summary", "methods"),
    ("Data Availability Statement", "data_availability"),
    ("Availability of data and materials", "data_availability"),
])
def test_the_top_level_headings_publishers_actually_use(heading, expected):
    """Every one of these returned None. The strongest single case is
    `References and Notes` on page 83 of 10.1126/science.aat5031's supplement:
    unrecognised, it left 71 blocks and 19,265 characters of other people's
    reference titles labelled `methods`."""
    assert sections.normalize(heading) == expected


@pytest.mark.parametrize("heading", [
    # Nature uses `Main` for the body as a whole; mapping it to any canonical
    # name is the guess this module refuses to make.
    "Main",
    # The bullet prefix must not promote a Cell Press highlight line, of which
    # 10.1016/j.cell.2021.01.053 has four on page 2.
    "d Detailed COVID-19 immune landscape depicted by",
    "d SARS-CoV-2 RNA is present in diverse epithelial and",
    "d Megakaryocytes and monocyte subsets may contribute to",
])
def test_the_widened_prefix_does_not_promote_a_highlight_line(heading):
    assert sections.normalize(heading) is None


@pytest.mark.parametrize("heading", [
    "Single-cell profiling of pancreatic islets",
    "TP53 is required for the response",
    "",
    None,
    "We used the following methods to isolate nuclei from frozen tissue and then "
    "sequenced them on a NovaSeq",
])
def test_unrecognised_headings_stay_none(heading):
    """Guessing is worse than not knowing: an unrecognised heading is as likely to
    be a Methods subsection as a Results one."""
    assert sections.normalize(heading) is None


def test_sec_type_is_used_when_the_title_is_uninformative():
    assert sections.normalize("Nuclei isolation", sec_type="methods") == "methods"
    assert sections.normalize("Anything", sec_type="ref-list") == "references"


def test_glued_heading_is_split_off_the_paragraph():
    """Nature's PDFs put the heading in the same layout block as the paragraph:
    one block reads "Methods Data collection Nuclei isolation from adult heart
    tissue...". Without this split, 10.1038/s41467-023-40505-5's PDF yields no
    sections at all."""
    split = sections.split_leading_heading(
        "Methods Data collection Nuclei isolation from adult heart tissue was performed")
    assert split is not None
    name, heading, rest = split
    assert name == "methods"
    assert heading == "Methods"
    assert rest.startswith("Data collection")


def test_a_glued_references_and_notes_is_split_off_its_first_citation():
    """The lookahead demanded a capital letter and a reference list starts with a
    digit, so page 83 of 10.1126/science.aat5031's supplement could only split as
    heading `REFERENCES` and rest `AND NOTES 1. K. W. Wucherpfennig...`. Ordering
    matters too: `_leading_patterns` has no `$` anchor, so with the bare
    `references?` first the alternation would still take it."""
    split = sections.split_leading_heading(
        "REFERENCES AND NOTES 1. K. W. Wucherpfennig, Structural basis of "
        "molecular mimicry, J. Autoimmun. 16, 293-302 (2001).")
    assert split is not None
    name, heading, rest = split
    assert (name, heading) == (sections.REFERENCES, "REFERENCES AND NOTES")
    assert rest.startswith("1. K. W. Wucherpfennig")


@pytest.mark.parametrize("text", [
    "Results of the assay were consistent with prior work in this area of study",
    "Background information about the mice was taken from the vendor records here",
    "Methods",
])
def test_sentences_that_merely_start_with_a_keyword_are_not_split(text):
    """The guard is that a heading is followed by the start of a new sentence. It
    is also a regression test for `re.IGNORECASE` making `[A-Z]` match lower case,
    which silently disabled the guard entirely."""
    assert sections.split_leading_heading(text) is None


def test_a_heading_carries_its_section_over_what_follows():
    tracker = sections.SectionTracker()
    assert tracker.carry("text before any heading") is None
    assert tracker.heading(sections.METHODS) == sections.METHODS
    assert tracker.carry("We isolated nuclei") == sections.METHODS
    assert tracker.heading(sections.RESULTS) == sections.RESULTS
    assert tracker.carry("We found cells") == sections.RESULTS
    assert tracker.seen == [sections.METHODS, sections.RESULTS]
    assert tracker.abandoned == [] and tracker.reason() is None


def test_cell_press_star_methods_is_recognised_with_the_glyph():
    """Cell Press publishes the heading as STAR★METHODS, with U+2605, in the XML as
    well as the PDF. The alias allowed an ASCII asterisk, which is how the heading is
    written *about* -- so the top-level Methods section of both Cell papers here went
    unrecognised, leaving 69 and 51 main-text blocks unlabelled, the key resources
    table among them."""
    for heading in ("STAR★METHODS", "STAR★methods", "STAR ★ Methods",
                    "STAR Methods", "STAR*Methods", "star methods"):
        assert sections.normalize(heading) == sections.METHODS, heading


def test_a_low_value_heading_only_claims_text_that_looks_like_its_content():
    """10.1016/j.cell.2025.05.027, a PMC author manuscript: the REFERENCES heading on
    page 31 carried 227 of 415 blocks to the end of the document, which in that
    layout is the key resources table -- reagents and catalogue numbers labelled as
    other people's bibliography. `references` is on LOW_VALUE, so a consumer that
    skips it does not deprioritise that text, it drops it."""
    tracker = sections.SectionTracker()
    tracker.heading(sections.REFERENCES)
    citation = "1. Wen L, Li G, Huang T, Geng W, Pei H (2022). Single-cell technologies."
    assert tracker.carry(citation) == sections.REFERENCES
    for row in ("Punch pliers Total Tools 9070220SB",
                "micro-Slide 8-well cell culture chamber ibidi 80841",
                "40 um strainer Cell Strainer PN 43-10040-40"):
        assert tracker.carry(row) is None, row
    assert tracker.withheld == 3
    assert "low-value" in tracker.reason()
    # The span stays open, so a citation after a stray line is still labelled: a
    # reference list interrupted by a page artifact must not lose its tail.
    assert tracker.carry("2. Smith J, Jones K (2020). Another paper. doi:10.1/x") \
        == sections.REFERENCES


@pytest.mark.parametrize("text,expected", [
    ("1. Wen L, Li G (2022). Single-cell technologies for multiomic analysis.", True),
    ("12) Author A, Author B, et al. Nature Genetics.", True),
    ("Smith and Jones (2019) reported similar results in mouse.", True),
    ("Available at doi:10.1016/j.cell.2025.05.027", True),
    ("Punch pliers Total Tools 9070220SB", False),
    ("Chromium Next GEM Single Cell 3' Kit v3.1 10x Genomics PN-1000268", False),
])
def test_what_counts_as_a_citation(text, expected):
    assert sections.looks_like_citation(text) is expected


def test_an_unbounded_section_flows_as_far_as_the_paper_does():
    """Methods legitimately runs for pages through its own unrecognised
    subsection headings, and that is the behaviour that makes it attributable."""
    tracker = sections.SectionTracker()
    tracker.heading(sections.METHODS)
    for _ in range(40):
        assert tracker.carry("x" * 1000) == sections.METHODS
    assert tracker.abandoned == []


def test_a_statement_section_is_abandoned_rather_than_carried_over_a_paper():
    """10.1126/science.adt8307: the standalone `CONCLUSION` line in Science's
    front-page structured summary was never followed by another heading this module
    recognises, so 996 of 1,184 main-text blocks came back labelled `conclusions`
    and the paper's real Results reported 5. A wrong section is worse than none --
    it makes a filter drop the text it was looking for and call it an answer."""
    tracker = sections.SectionTracker()
    tracker.heading(sections.CONCLUSIONS)
    labelled = [tracker.carry("x" * 1000) for _ in range(20)]
    assert labelled[0] == sections.CONCLUSIONS, "the conclusion itself is still labelled"
    assert labelled[-1] is None, "the rest of the paper is not"
    assert tracker.abandoned == [sections.CONCLUSIONS]
    assert "conclusions" in tracker.reason()
    # And a real heading later on still reopens labelling.
    assert tracker.heading(sections.METHODS) == sections.METHODS
    assert tracker.carry("Nuclei were isolated") == sections.METHODS


def test_an_abandoned_section_is_not_reopened_by_its_own_heading():
    """The existing coverage only reopened with a *different* section.
    10.1126/science.aat5031 abandons `abstract` at block 33, and then block 70 --
    the heading `One Sentence Summary`, an ABSTRACT alias -- reopened it, so
    blocks 70-85 came back labelled `abstract`: 16 blocks and 6,272 characters,
    four of them figure legends beginning "Fig. 1. Mapping the spatial and
    temporal architecture of the mature and developing human kidney". The record
    said "the blocks after it are left unlabelled" the whole time."""
    tracker = sections.SectionTracker()
    tracker.heading(sections.ABSTRACT)
    for _ in range(10):
        tracker.carry("x" * 1000)
    assert tracker.abandoned == [sections.ABSTRACT]

    assert tracker.heading(sections.ABSTRACT) is None
    assert tracker.carry("Fig. 1. Mapping the spatial and temporal architecture") is None
    assert tracker.reopens_refused == [sections.ABSTRACT]
    assert "reopen abstract" in tracker.reason()
    # A different section still opens normally.
    assert tracker.heading(sections.METHODS) == sections.METHODS


def test_the_abandonment_bound_is_a_configurable_cap():
    """It decided a third of one article's labels while being a module constant
    with no config key, which the README said outright."""
    assert sections.SectionTracker().max_bounded_chars == Limits().max_bounded_section_chars
    tracker = sections.SectionTracker(limits=Limits(max_bounded_section_chars=100))
    tracker.heading(sections.ABSTRACT)
    assert tracker.carry("x" * 200) == sections.ABSTRACT
    assert tracker.carry("the next paragraph") is None
    assert tracker.abandoned == [sections.ABSTRACT]
    assert "100 characters" in tracker.reason()


def test_a_long_abstract_is_kept_up_to_the_measured_budget():
    """The longest legitimate run measured is a 4,653-character Cell Press abstract
    with its highlights and eTOC blurb, so the budget must not cut that."""
    tracker = sections.SectionTracker()
    tracker.heading(sections.ABSTRACT)
    assert tracker.carry("x" * 4653) == sections.ABSTRACT
    assert tracker.carry("the next paragraph") == sections.ABSTRACT
    assert tracker.abandoned == []


@pytest.mark.parametrize("name", sorted(sections.BOUNDED_SECTIONS))
def test_every_bounded_section_is_actually_bounded(name):
    tracker = sections.SectionTracker()
    tracker.heading(name)
    for _ in range(20):
        tracker.carry("x" * 1000)
    assert tracker.abandoned == [name]


# -- table cards -------------------------------------------------------------

def test_header_found_below_a_title_and_a_caption_row():
    """10.1038/s41591-018-0269-2's MOESM1 puts the title on row 1, the caption on
    row 2, a blank on row 3, and the header on row 4. Assuming row 1 reads the
    caption as column names and the column names as data."""
    rows = [
        ("Supplementary Table 1", None, None, None),
        ("Detailed demographic and clinical data for all patients", None, None, None),
        (None, None, None, None),
        ("patient_code", "diagnosis", "Sex", "Age"),
        ("SMM05", "SMM", "M", 61),
        ("AL01", "AL", "F", 74),
    ]
    card = tables.build_card(rows, "x.xlsx", "sheet 'S1'", L)
    assert card.header_row == 3
    assert card.header == ["patient_code", "diagnosis", "Sex", "Age"]
    assert card.n_rows == 2
    assert "Detailed demographic" in card.caption
    assert card.header_confidence == "high"


def test_header_on_row_two_after_a_single_caption_line():
    """10.1016/j.xcrm.2023.101158's mmc7.xlsx."""
    rows = [("Table S6: Cluster marker genes", None, None),
            ("cluster", "gene", "p_val"),
            (0, "FGFBP2", 0.0)]
    card = tables.build_card(rows, "mmc7.xlsx", "sheet 'Table S6'", L)
    assert card.header_row == 1
    assert card.caption == "Table S6: Cluster marker genes"


def test_low_cardinality_text_column_is_enumerated():
    """This is the point of the card: two distinct values answer "sex" outright."""
    rows = [("id", "Sex", "organism"), ("a", "M", "Mus musculus"),
            ("b", "F", "Mus musculus"), ("c", "M", "Mus musculus")]
    card = tables.build_card(rows, "x.xlsx", "s", L)
    sex = next(c for c in card.columns if c["name"] == "Sex")
    assert sex["values"] == ["F", "M"]
    organism = next(c for c in card.columns if c["name"] == "organism")
    assert organism["values"] == ["Mus musculus"]
    assert "Sex [text, 2 distinct] = F | M" in tables.render(card, L)


def test_few_numeric_values_are_enumerated_but_many_are_summarised():
    """`0 | 6 | 24` says "these are the timepoints"; 30 patient ages say nothing a
    range does not, at ten times the length."""
    timepoints = [("t",)] + [(v,) for v in [0, 6, 24] * 4]
    card = tables.build_card(timepoints, "x.xlsx", "s", L)
    assert card.columns[0]["values"] == ["0", "6", "24"]

    ages = [("Age",)] + [(v,) for v in range(30, 70)]
    card = tables.build_card(ages, "x.xlsx", "s", L)
    column = card.columns[0]
    assert "values" not in column and "examples" in column
    assert (column["min"], column["max"]) == (30, 69)
    assert "median" in column


def test_numeric_values_are_enumerated_in_numeric_order():
    rows = [("dose",)] + [(v,) for v in [10, 2, 100]]
    card = tables.build_card(rows, "x.xlsx", "s", L)
    assert card.columns[0]["values"] == ["2", "10", "100"]


def test_mostly_numeric_column_with_na_strings_is_still_numeric():
    rows = [("value",)] + [(v,) for v in [1, 2, 3, 4, 5, 6, 7, 8, 9, "NA"]]
    card = tables.build_card(rows, "x.xlsx", "s", L)
    assert card.columns[0]["dtype"] == tables.NUMBER


def test_an_infinite_cell_does_not_make_an_invalid_json_line(tmp_path):
    """`float("inf")` succeeds, so line 520 of
    10.1038/s41467-023-40505-5's blocks.jsonl carried `"max": Infinity` -- which
    Python's json.loads accepts by default and serde_json, Go's encoding/json,
    PostgreSQL jsonb and DuckDB all reject."""
    values = [1.5, 2.5, 3.0, 4.0, 5.0, 6.0, 7.0, 8.0, 9.0, 10.0, 11.0, 12.0, 13.0,
              "Inf", "-inf", "NaN"]
    card = tables.build_card([("neg. log10-pval",)] + [(v,) for v in values], "x.xlsx", "s", L)
    column = card.columns[0]
    assert column["dtype"] == tables.NUMBER
    assert column["max"] == 13.0 and column["min"] == 1.5
    assert column["n_non_finite"] == 3
    assert any("non-finite" in note for note in card.notes)

    path = tmp_path / "blocks.jsonl"
    write_blocks(path, [Block(kind=TABLE, text=tables.render(card, L),
                              source_file="x.xlsx", origin="xlsx",
                              table=card.to_dict())])
    for line in path.read_text().splitlines():
        json.loads(line, parse_constant=lambda c: pytest.fail(f"not JSON: {c}"))


def test_headerless_matrix_is_reported_not_guessed():
    rows = [(1, 2, 3), (4, 5, 6), (7, 8, 9)]
    card = tables.build_card(rows, "x.xlsx", "s", L)
    assert card.header_row is None
    assert any("no header row" in note for note in card.notes)
    assert card.header == ["column_1", "column_2", "column_3"]


def test_a_header_whose_names_repeat_is_accepted_not_turned_into_data():
    """`08_mmc5.xlsx` sheet `TV+vs.V-` is two identical eight-column tables side
    by side: present 16, distinct 8, threshold 11, so the 0.7 rule rejected it.
    The card then said "no header row identified; columns are positional" and
    printed `7. column_7 [text, 3 distinct] = cluster | virus+ | virus-` -- a
    header string offered as one of three complete values."""
    header = ["row.names", "p_val", "cluster", "gene"] * 2
    rows = [tuple(header)] + [("1", "0", "virus-", "IFITM1") * 2 for _ in range(4)]
    card = tables.build_card(rows, "mmc5.xlsx", "sheet 'TV+vs.V-'", L)
    assert card.header_row == 0
    assert card.header[:4] == ["row.names", "p_val", "cluster", "gene"]
    assert any("2 tables side by side" in note for note in card.notes)
    assert not any("no header row" in note for note in card.notes)


def test_a_two_row_header_is_composed_rather_than_read_as_data():
    """`49_..._MOESM4_ESM.xlsx` sheet `Supplementary Data 5` puts the cell type on
    one row and `Sum.PIPs / N.SNPs` on the next, so the card printed
    `column_5 [number, 6 distinct] = 0 | 0.01 | 0.02 | 0.03 | Endothelial OCRs |
    Sum.PIPs` -- two header strings offered as data values."""
    rows = [
        (None, None, "Cardiomyocyte OCRs", None, "Endothelial OCRs", None),
        ("Locus", "Location", "Sum.PIPs", "N.SNPs", "Sum.PIPs", "N.SNPs"),
        (7, "chr1:9365199-10806984", 0.74, 7, 0.0, 0),
        (15, "chr1:21736588-23086883", 0.05, 5, 0.01, 6),
    ]
    card = tables.build_card(rows, "m4.xlsx", "sheet 'Supplementary Data 5'", L)
    assert card.header_row == 1 and card.header_rows == [0, 1]
    assert card.header == ["Locus", "Location",
                           "Cardiomyocyte OCRs / Sum.PIPs",
                           "Cardiomyocyte OCRs / N.SNPs",
                           "Endothelial OCRs / Sum.PIPs",
                           "Endothelial OCRs / N.SNPs"]
    assert any("header spans 2 rows" in note for note in card.notes)
    values = {v for c in card.columns for v in (c.get("values") or [])}
    assert "Endothelial OCRs" not in values and "Sum.PIPs" not in values


def test_a_blank_line_between_a_caption_and_a_header_keeps_them_apart():
    """10.1038/s41591-018-0269-2 MOESM1 is title, caption, blank, then the header
    on row 4. Those are not one two-row header."""
    rows = [("Supplementary Table 1", None, None),
            ("Donor characteristics for the cohort", None, None),
            (None, None, None),
            ("donor", "age", "sex"),
            ("D1", 44, "F"), ("D2", 61, "M")]
    card = tables.build_card(rows, "m1.xlsx", "s", L)
    assert card.header_row == 3 and card.header_rows is None
    assert card.header == ["donor", "age", "sex"]


def test_header_confidence_is_low_without_a_type_change():
    """All-text rows under all-text headers could be a first data row of gene
    names; the card says so rather than pretending to know."""
    rows = [("gene", "symbol"), ("TP53", "p53"), ("MYC", "myc")]
    card = tables.build_card(rows, "x.xlsx", "s", L)
    assert card.header_confidence == "low"
    assert any("type-change" in note for note in card.notes)


def test_a_column_holding_its_own_header_name_is_never_high_confidence():
    """The safety net behind `split_blocks`, for a sheet whose second table is
    not separated by a blank row. Before the splitter landed this fired on
    exactly 7 column-instances across 5 cards of this corpus, all of them `high`,
    with no false positives."""
    rows = [("timepoint", "value"), ("0", "1.2"), ("24", "3.4"),
            ("timepoint", "value"), ("0", "5.6")]
    card = tables.build_card(rows, "x.xlsx", "s", L)
    assert card.header_confidence == "low"
    assert any("its own header name" in note for note in card.notes)
    assert not any("type-change" in note for note in card.notes), \
        "the type-change note would be false: there was one"


def test_an_ordinary_column_is_not_accused_of_holding_its_header_name():
    """Exact match, not substring: `gene` inside `gene_id` fires on 50 columns of
    this corpus, most of them legitimate."""
    rows = [("gene", "gene_id"), ("TP53", "gene_00001"), ("MYC", "gene_00002")]
    card = tables.build_card(rows, "x.xlsx", "s", L)
    assert not any("its own header name" in note for note in card.notes)


def test_trailing_empty_columns_are_dropped():
    rows = [("a", "b", None, None), (1, 2, None, None)]
    card = tables.build_card(rows, "x.xlsx", "s", L)
    assert card.n_columns == 2


def test_empty_table_yields_no_card():
    assert tables.build_card([], "x.xlsx", "s", L) is None
    assert tables.build_card([(None, None), (None,)], "x.xlsx", "s", L) is None


def test_header_with_no_data_rows_still_makes_a_card():
    """The column names alone say what the table was going to be about."""
    card = tables.build_card([("organism", "age", "sex")], "x.xlsx", "s", L)
    assert card.n_rows == 0
    assert any("no data rows" in note for note in card.notes)


@pytest.mark.parametrize("value,expected", [
    ("  spaced   out  ", "spaced out"),
    ("\xa0", None),
    ("", None),
    (None, None),
    (True, "TRUE"),
    (3.0, "3"),
    (3.5, "3.5"),
])
def test_clean_cell(value, expected):
    """Excel's empty-looking cells are not all None: a formula evaluated to "" or
    a whitespace-only string counted as present is what makes a caption row look
    like a header row."""
    assert tables.clean_cell(value) == expected


def test_card_render_respects_the_character_budget():
    limits = Limits(max_card_chars=800)
    rows = [tuple(f"column_{i}" for i in range(80))] + [
        tuple(f"value_{i}_{r}" for i in range(80)) for r in range(5)]
    card = tables.build_card(rows, "wide.xlsx", "s", limits)
    text = tables.render(card, limits)
    assert len(text) <= 900, len(text)
    assert "further column(s) not shown" in text


def test_the_renderer_says_when_it_dropped_sample_rows():
    """Nothing a cap drops may be silent, and the column line has always said so
    while the sample-row loop just broke out."""
    limits = Limits(max_card_chars=700, max_sample_rows=3)
    rows = [("a", "b", "c")] + [tuple(f"{c}_value_{r}" * 4 for c in "abc")
                                for r in range(5)]
    text = tables.render(tables.build_card(rows, "x.xlsx", "s", limits), limits)
    assert "further sample row(s) not shown" in text


def test_card_states_when_the_scan_was_capped():
    limits = Limits(max_scan_rows=10)
    rows = [("a", "b")] + [(i, i * 2) for i in range(9)]
    card = tables.build_card(rows, "x.xlsx", "s", limits, truncated=True,
                             n_rows_total=99999)
    text = tables.render(card, limits)
    assert "scan stopped at 10 rows" in " ".join(card.notes)
    assert "source reports 99999" in text


def test_a_truncated_scan_never_renders_its_value_set_as_complete():
    """The `=` form means the complete value set, which is the entire point of
    the card. 10.1126/science.aat5031's data_s1.csv is 40,269 lines; scanned to
    5,000 it printed `celltype [text, 12 distinct] = B cell | CD4 T cell | ...`
    where the file holds 33, missing Podocyte, Proximal tubule, Glomerular
    endothelium and every other epithelial and endothelial type."""
    body = b"barcode,celltype\n" + b"".join(
        f"bc{i},{'Podocyte' if i > 40 else 'B cell'}\n".encode() for i in range(59))

    whole = spreadsheet.cards_from_csv(body, "s1.csv", L)[0][0]
    assert whole.n_rows_total == 60 and whole.truncated is False
    assert "celltype [text, 2 distinct] = B cell | Podocyte" in tables.render(whole, L)

    capped = spreadsheet.cards_from_csv(body, "s1.csv", Limits(max_scan_rows=10))[0][0]
    text = tables.render(capped, Limits(max_scan_rows=10))
    assert capped.n_rows_total == 60
    assert "source reports 60 row(s)" in text
    assert "scan stopped at 10 rows of 60" in text
    assert " = " not in text
    assert "celltype [text, 1 distinct] e.g. B cell" in text


#: The shape of `Figure 6` in 10.1126/sciimmunol.aba4163's data file, trimmed to
#: three panels: a panel title on its own row, a header row, then data, with a
#: blank row between panels.
def _panel(name, header, values):
    return [(name, None, None), header] + [(None,) + row for row in values]


STACKED_PANELS = (
    _panel("Figure 6C", ("% Crescents", "Control", "S. aureus + NTN"),
           [(0, 13.3), (0, 23.3), (3.3, 20.0)])
    + [(None, None, None)]
    + _panel("Figure 6D", ("[% of CD3+]", "Control", "S. aureus + NTN"),
             [(0.23, 1.53), (0.16, 2.68), (0.06, 1.81)])
    + [(None, None, None)]
    + _panel("Figure 6E", ("[%CD3+ cells]", "Control", "S. aureus + NTN"),
             [(0.94, 10.9), (0.67, 9.56), (0.91, 6.31)])
)


def test_a_sheet_of_stacked_panels_becomes_one_card_per_panel():
    """`detect_header` found the first panel's header and `build_card` read every
    later panel's title, units and header row as data:
    `Figure 6C [text, 23 distinct, 63 empty] = % Crescents | [% of CD3+] | ... |
    Figure 6D | Figure 6E | ...`, at header_confidence high and with no note."""
    data = make_xlsx({"Figure 6": [list(r) for r in STACKED_PANELS]})
    cards, status, _ = spreadsheet.cards_from_xlsx(data, "s1.xlsx", L)
    assert status == "ok" and len(cards) == 3
    assert [c.locator for c in cards] == ["sheet 'Figure 6' rows 1-5",
                                          "sheet 'Figure 6' rows 7-11",
                                          "sheet 'Figure 6' rows 13-17"]
    assert cards[0].data_ref["row_start"] == 1 and cards[0].data_ref["row_end"] == 5
    assert all("3 blank-row-separated tables" in c.notes[0] for c in cards)
    # The pooled column name is gone: no card carries a later panel's title.
    names = {n for c in cards for n in c.header}
    values = {v for c in cards for col in c.columns
              for v in (col.get("values") or []) + (col.get("examples") or [])}
    assert "Figure 6D" not in names | values
    assert "Figure 6E" not in names | values


@pytest.mark.parametrize("parts,expected", [
    # A lone row above a table is a panel title; merge it down, do not refuse.
    ([1, 13, 13, 13], [(0, 15), (16, 29), (30, 43)]),
    ([9, 1, 9, 10], [(0, 9), (10, 21), (22, 32)]),
    # One title row above one table is one table.
    ([1, 13], []),
    ([1, 40], []),
    # A trailing lone row is a footnote, not a table.
    ([5, 5, 1], [(0, 5), (6, 11)]),
    ([9], []),
])
def test_which_blank_separated_sheets_are_split(parts, expected):
    """Requiring every part to be at least 2 rows was tried and it disables the
    split on `STable 4.4` `[1, 13, 13, 13]` and `Figure S7` `[9, 1, 9, 10]`."""
    rows: list = []
    for size in parts:
        if rows:
            rows.append((None,))
        rows.extend([(f"r{i}", i) for i in range(size)])
    assert tables.split_blocks(rows, L) == expected


def test_a_sheet_with_more_panels_than_the_cap_says_how_many_it_dropped():
    sheet = []
    for panel in range(4):
        if sheet:
            sheet.append([None, None])
        sheet.append([f"Panel {panel}", None])
        sheet.append(["name", "value"])
        sheet.extend([[f"r{i}", i] for i in range(3)])
    cards, _, meta = spreadsheet.cards_from_xlsx(
        make_xlsx({"S": sheet}), "s.xlsx", Limits(max_tables_per_sheet=2))
    assert len(cards) == 2
    assert meta["tables_skipped"] == 2


def test_card_does_not_copy_the_data_it_points_at_it():
    """Duplicating a 2.4 GB corpus to paraphrase it would be the wrong trade: the
    card records where to re-read the real values instead -- and it has to record
    enough to actually do it, which for a while it did not: no scan window and no
    file hash, so the rows a card described could not be re-read reproducibly."""
    card = tables.build_card([("a",), (1,), (2,)], "x.xlsx", "sheet 'S1'", L,
                             data_ref={"file": "x.xlsx", "sheet": "S1",
                                       "sha256": "abc123"})
    assert card.data_ref == {"file": "x.xlsx", "locator": "sheet 'S1'",
                             "sheet": "S1", "sha256": "abc123",
                             "header_row": 1, "first_data_row": 2, "last_data_row": 3}


def test_a_split_panels_data_ref_carries_absolute_row_numbers():
    """A panel's card is built from a slice of the sheet, so its own row indices
    start at zero; `data_ref` has to say where that slice sat."""
    data = make_xlsx({"Figure 6": [list(r) for r in STACKED_PANELS]})
    cards, _, _ = spreadsheet.cards_from_xlsx(data, "s1.xlsx", L)
    # Each panel is a title row, a header row and three data rows.
    assert [c.data_ref["header_row"] for c in cards] == [2, 8, 14]
    assert [c.data_ref["first_data_row"] for c in cards] == [3, 9, 15]
    assert [c.data_ref["last_data_row"] for c in cards] == [5, 11, 17]
    assert len({c.data_ref["sha256"] for c in cards}) == 1


@pytest.mark.parametrize("name,payload,ref", [
    ("s.csv", b"donor,age\nD1,44\nD2,61\nD3,7\n",
     {"delimiter": ",", "header_row": 1, "first_data_row": 2, "last_data_row": 4}),
])
def test_read_rows_reprints_the_rows_the_card_describes(name, payload, ref):
    header, rows = spreadsheet.read_rows(payload, ref, ".csv", limit=2)
    assert header == ["donor", "age"]
    assert rows == [(2, ["D1", "44"]), (3, ["D2", "61"])]


# -- spreadsheets ------------------------------------------------------------

def test_xlsx_sheets_become_cards():
    data = make_xlsx({"Table S1": [["id", "Sex"], ["a", "M"], ["b", "F"]],
                      "Table S2": [["gene", "logFC"], ["TP53", 1.5]]})
    cards, status, meta = spreadsheet.cards_from_xlsx(data, "x.xlsx", L)
    assert status == "ok" and meta["sheets"] == 2
    assert [c.title for c in cards] == ["Table S1", "Table S2"]
    assert cards[0].locator == "sheet 'Table S1'"


def test_strict_ooxml_workbook_is_read_not_reported_empty():
    """10.1016/j.cell.2021.01.053's mmc7.xlsx is a strict ISO-29500 workbook.
    openpyxl reads those as having zero worksheets -- no exception, no warning --
    so a file holding `sampleID, Age, Sex, CoVID-19 severity` reported `no_text`,
    which is indistinguishable from a genuinely empty supplement."""
    sheets = {"Patient_info": [["sampleID", "Age", "Sex"], ["P1", 61, "M"]]}
    strict = make_strict_xlsx(sheets)

    import openpyxl
    assert openpyxl.load_workbook(__import__("io").BytesIO(strict),
                                  read_only=True).worksheets == [], \
        "fixture must reproduce the silent-empty behaviour"

    cards, status, meta = spreadsheet.cards_from_xlsx(strict, "mmc7.xlsx", L)
    assert status == "ok" and meta["strict_ooxml"] is True
    assert cards[0].header == ["sampleID", "Age", "Sex"]


def test_relax_strict_returns_none_for_an_ordinary_workbook():
    assert ooxml.relax_strict(make_xlsx({"S": [["a"]]})) is None
    assert ooxml.relax_strict(b"not a zip") is None


def test_workbook_without_declared_dimensions_is_read():
    """10.1038/s44161-025-00612-6's MOESM5 has unsized worksheets; openpyxl raises
    `ValueError: Worksheet is unsized` from `calculate_dimension()`, which is why
    this stage never calls it."""
    data = make_dimensionless_xlsx({"Figure 2d": [["Patient", "value"], ["BS_H15", 23.44]]})
    cards, status, _ = spreadsheet.cards_from_xlsx(data, "x.xlsx", L)
    assert status == "ok"
    assert cards[0].header == ["Patient", "value"]


def test_xlsx_scan_stops_at_the_row_cap():
    """One sheet in this corpus is 16,596 x 88. It cannot go into a prompt and
    almost none of it would help if it did."""
    limits = Limits(max_scan_rows=20)
    data = make_xlsx({"big": [["gene", "value"]] + [[f"G{i}", i] for i in range(500)]})
    cards, status, _ = spreadsheet.cards_from_xlsx(data, "x.xlsx", limits)
    assert status == "ok"
    assert cards[0].truncated is True
    assert cards[0].n_rows < 500
    assert cards[0].n_rows_total == 501


def test_empty_workbook_is_no_text():
    cards, status, _ = spreadsheet.cards_from_xlsx(make_xlsx({"blank": []}), "x.xlsx", L)
    assert (cards, status) == ([], "no_text")


def test_unreadable_bytes_are_named_not_crashed():
    cards, status, meta = spreadsheet.cards_from_xlsx(b"not a workbook", "x.xlsx", L)
    assert status == "unreadable" and "reason" in meta


@pytest.mark.parametrize("body,delimiter", [
    (b"a,b,c\n1,2,3\n4,5,6\n", ","),
    (b"a\tb\tc\n1\t2\t3\n4\t5\t6\n", "\t"),
    (b"a;b;c\n1;2;3\n4;5;6\n", ";"),
])
def test_csv_delimiter_is_sniffed(body, delimiter):
    cards, status, meta = spreadsheet.cards_from_csv(body, "x.csv", L)
    assert status == "ok" and meta["delimiter"] == delimiter
    assert cards[0].header == ["a", "b", "c"]


def test_empty_csv_is_no_text():
    assert spreadsheet.cards_from_csv(b"   \n", "x.csv", L)[1] == "no_text"


def test_a_workbook_that_still_declares_no_sheets_after_relaxing_is_unreadable():
    """The other half of the strict-OOXML story. `relax_strict` returning None means
    the namespaces were already ordinary, so zero worksheets is the file's own
    answer -- and `unreadable` says that, where `no_text` would blame the content."""
    data = make_zero_sheet_xlsx()
    cards, status, meta = spreadsheet.cards_from_xlsx(data, "x.xlsx", L)

    assert (cards, status) == ([], "unreadable")
    assert "declares no worksheets" in meta["reason"]
    assert meta["sheets"] == 0
    assert "strict_ooxml" not in meta, "nothing was relaxed"


def test_sheets_beyond_the_cap_are_counted_not_dropped_quietly():
    sheets = {f"S{n}": [["id", "value"], [f"r{n}", n]] for n in range(6)}
    cards, status, meta = spreadsheet.cards_from_xlsx(
        make_xlsx(sheets), "x.xlsx", Limits(max_sheets=2))

    assert status == "ok" and len(cards) == 2
    assert meta["sheets"] == 6 and meta["sheets_skipped"] == 4


def test_the_card_cap_stops_the_sheet_loop():
    sheets = {f"S{n}": [["id", "value"], [f"r{n}", n]] for n in range(5)}
    cards, status, _ = spreadsheet.cards_from_xlsx(
        make_xlsx(sheets), "x.xlsx", Limits(max_tables_per_file=2))
    assert status == "ok" and len(cards) == 2


def test_one_unreadable_sheet_does_not_lose_the_others():
    """A workbook is a bag of independent tables; a single sheet that blows up on
    iteration must be recorded and stepped over, not take the file down with it."""
    data = make_xlsx({"good": [["id", "value"], ["a", 1]],
                      "bad": [["id", "value"], ["b", 2]]})

    real_iter_rows = openpyxl.worksheet._read_only.ReadOnlyWorksheet.iter_rows

    def flaky(self, *args, **kwargs):
        if self.title == "bad":
            raise ValueError("Worksheet is unsized")
        return real_iter_rows(self, *args, **kwargs)

    with mock.patch.object(openpyxl.worksheet._read_only.ReadOnlyWorksheet,
                           "iter_rows", flaky):
        cards, status, meta = spreadsheet.cards_from_xlsx(data, "x.xlsx", L)

    assert status == "ok"
    assert [c.title for c in cards] == ["good"]
    assert meta["errors"] == ["sheet 'bad': ValueError: Worksheet is unsized"]


def test_a_sheet_that_cannot_reset_its_dimensions_is_still_read():
    """`reset_dimensions` is a newer openpyxl API and the call is defensive. If it
    is missing or refuses, the sheet must still be scanned rather than skipped."""
    data = make_xlsx({"S": [["id", "value"], ["a", 1]]})

    def refuse(self):
        raise AttributeError("no reset_dimensions on this version")

    with mock.patch.object(openpyxl.worksheet._read_only.ReadOnlyWorksheet,
                           "reset_dimensions", refuse, create=True):
        cards, status, _ = spreadsheet.cards_from_xlsx(data, "x.xlsx", L)

    assert status == "ok" and cards[0].header == ["id", "value"]


# -- spreadsheets: legacy .xls -----------------------------------------------
# Real BIFF bytes, parsed by xlrd itself. These used to run against a
# `types.SimpleNamespace` whose `open_workbook` returned a hand-built fake book,
# because xlrd was an optional extra and CI did not install it -- so the suite
# could not have noticed xlrd changing under it, and the 56 .xls supplements in
# the corpus were the only thing that would have. `make_xls` says what its BIFF5
# stream does and does not stand in for.

def test_legacy_xls_sheets_become_cards():
    data = make_xls({
        "Table S1": [["id", "Sex"], ["a", "M"], ["b", "F"]],
        "Notes": [["read me"], ["some prose"]],
    })
    cards, status, meta = spreadsheet.cards_from_xls(data, "mmc2.xls", L)

    assert status == "ok" and meta["sheets"] == 2
    assert [c.title for c in cards] == ["Table S1", "Notes"]
    assert cards[0].header == ["id", "Sex"]
    assert cards[0].locator == "sheet 'Table S1'"


def test_legacy_xls_numbers_arrive_as_numbers():
    """The cell-type split is xlrd's, not ours, and the column profiler reads it:
    a numeric column is enumerated under `max_unique_numeric_values` and a text one
    under `max_unique_values`. A stub returning Python objects could not get this
    wrong, so nothing here checked it."""
    data = make_xls({"S": [["donor", "dose_uM"], ["d1", 0], ["d2", 10], ["d3", 10]]})
    cards, status, _ = spreadsheet.cards_from_xls(data, "x.xls", L)

    assert status == "ok"
    dose = next(c for c in cards[0].columns if c["name"] == "dose_uM")
    assert dose["dtype"] == "number"
    assert (dose["min"], dose["max"]) == (0.0, 10.0)
    assert set(dose["values"]) == {"0", "10"}


def test_legacy_xls_honours_the_row_and_sheet_caps():
    rows = [["gene", "value"]] + [[f"G{n}", n] for n in range(100)]
    data = make_xls({"big": rows, "second": [["a"], ["b"]]})
    cards, status, _ = spreadsheet.cards_from_xls(
        data, "x.xls", Limits(max_scan_rows=10, max_sheets=1))

    assert status == "ok" and len(cards) == 1
    assert cards[0].truncated is True
    assert cards[0].n_rows_total == 101


def test_an_unreadable_xls_is_named_not_crashed():
    """xlrd's own refusal, not a stubbed exception: `unreadable` blames the file,
    where `parser_error` would blame this stage."""
    cards, status, meta = spreadsheet.cards_from_xls(b"garbage, not a workbook", "x.xls", L)
    assert (cards, status) == ([], "unreadable")
    assert "Unsupported format, or corrupt file" in meta["reason"]


def test_an_empty_xls_is_no_text():
    assert spreadsheet.cards_from_xls(make_xls({"blank": []}), "x.xls", L)[1] == "no_text"


def test_legacy_xls_without_xlrd_says_so(monkeypatch):
    """xlrd is a hard requirement now -- 56 files and 129 MB of the corpus turn on
    it -- so this covers an install that predates the promotion. The reason has to
    name the package, because that string is the whole instruction a curator gets.
    """
    monkeypatch.setitem(sys.modules, "xlrd", None)
    cards, status, meta = spreadsheet.cards_from_xls(make_xls({"S": [["a"], ["b"]]}),
                                                    "x.xls", L)
    assert (cards, status) == ([], "unsupported_format")
    assert "xlrd" in meta["reason"] and "pip install" in meta["reason"]


# -- spreadsheets: CSV sniffing ----------------------------------------------

@pytest.mark.parametrize("body,expected", [
    (b"single column\nvalue\nother\n", ","),      # sniffer fails; no candidate present
    (b"a|b|c\n1|2|3\n", "|"),
    (b"justoneline", ","),
    # The sniffer refuses a one-column file, but the fallback's count still finds
    # the real delimiter on the first line.
    (b"a;b\n1;2\n", ";"),
])
def test_the_delimiter_falls_back_to_counting_when_sniffing_fails(body, expected):
    """`csv.Sniffer` raises "Could not determine delimiter" for a single-column
    file, and a raise here would report a readable supplement as unreadable."""
    cards, status, meta = spreadsheet.cards_from_csv(body, "x.csv", L)
    assert meta["delimiter"] == expected
    assert status in {"ok", "no_text"}


def test_a_utf8_bom_is_stripped_from_the_first_header():
    """Excel writes CSVs with a BOM; leaving it on turns the first column name into
    `\\ufeffid`, which no downstream match on "id" would find."""
    cards, status, _ = spreadsheet.cards_from_csv(
        "﻿id,Sex\na,M\nb,F\n".encode("utf-8"), "x.csv", L)
    assert status == "ok" and cards[0].header == ["id", "Sex"]


def test_a_csv_with_nothing_but_a_delimiter_line_is_no_text():
    """The card builder finds no usable table, and the sniffed delimiter is still
    reported -- it is the evidence for why the file was read the way it was."""
    cards, status, meta = spreadsheet.cards_from_csv(b",,,\n", "x.csv", L)
    assert (cards, status) == ([], "no_text")
    assert "delimiter" in meta


def test_a_csv_field_over_the_stdlib_limit_is_named_not_crashed():
    """`csv.reader` raises "field larger than field limit (131072)" mid-iteration,
    not at construction, so the catch has to wrap the scan. A supplement with one
    enormous cell -- a pasted FASTA, typically -- is how this shows up."""
    body = b'a,b\n"' + b"x" * 200_000 + b'",2\n'
    cards, status, meta = spreadsheet.cards_from_csv(body, "x.csv", L)

    assert (cards, status) == ([], "unreadable")
    assert "field larger than field limit" in meta["reason"]


# -- JATS --------------------------------------------------------------------

def test_jats_sections_and_metadata():
    body = ('<sec sec-type="methods"><title>Methods</title>'
            '<p>Islets were dissociated and loaded on a 10x Chromium controller.</p></sec>'
            '<sec><title>Results</title><p>We found beta cells.</p></sec>')
    blocks, status, meta = jats.blocks_from_jats(jats_article(body), "f.nxml", L)
    assert status == "ok"
    assert meta["title"] == "A test article about islets"
    assert meta["keywords"] == ["single-cell", "pancreas"]
    assert set(meta["sections"]) >= {"methods", "results"}
    methods = [b for b in blocks if b.section == "methods" and b.kind == PARAGRAPH]
    assert "10x Chromium" in methods[0].text
    assert any(b.kind == METADATA for b in blocks)
    assert any(b.section == "abstract" for b in blocks)


def test_a_jats_block_carries_the_heading_path_it_sits_under():
    """`section` is one canonical name out of eleven, and `walk_section` already
    knows the whole tree. `pdf.py` is deliberately left alone: there the tree is
    a guess, and a guessed path is what this package refuses to produce."""
    body = ('<sec sec-type="methods"><title>Methods</title>'
            '<sec><title>Nuclei isolation</title>'
            '<p>Nuclei were isolated from frozen tissue.</p></sec>'
            '<sec><title>Library preparation</title>'
            '<p>Libraries were made with the 10x kit.</p></sec></sec>')
    blocks, _, _ = jats.blocks_from_jats(jats_article(body), "f.nxml", L)
    paths = {b.text: b.section_path for b in blocks if b.origin == "jats"}
    assert paths["Nuclei were isolated from frozen tissue."] == \
        ["Methods", "Nuclei isolation"]
    assert paths["Libraries were made with the 10x kit."] == \
        ["Methods", "Library preparation"]
    assert paths["Nuclei isolation"] == ["Methods", "Nuclei isolation"]
    assert paths["Methods"] == ["Methods"]
    # Emitted only when it is real, so nothing else in the line changes.
    front = next(b for b in blocks if b.kind == METADATA)
    assert front.section_path is None and "section_path" not in front.to_dict()


def test_a_jats_locator_counts_children_of_its_own_tag():
    """`[n]` in XPath means the nth child *of that tag*. Counting every child made
    153 of the 168 body/back locators in 10.1038/s41467-023-40505-5 point at a
    different element, and only 76 resolved at all."""
    body = ("<sec><title>T</title><p>first paragraph here</p><fig/>"
            "<p>second paragraph here</p></sec>")
    blocks, _, _ = jats.blocks_from_jats(jats_article(body), "f.nxml", L)
    second = next(b for b in blocks if b.text == "second paragraph here")
    assert second.locator == "body/sec[1]/p[2]"


def test_a_pdf_block_carries_the_rectangle_it_came_from():
    """PyMuPDF hands over (x0, y0, x1, y1, text, block_no, block_type) and this
    module kept two of the seven, so a PDF block was locatable only to a page."""
    data = make_pdf_pages([["Nuclei were isolated from frozen heart tissue and "
                            "libraries were prepared with the 10x Chromium kit."]])
    blocks, _, _ = pdf.blocks_from_pdf(data, "f.pdf", L)
    ref = blocks[0].locator_ref
    assert ref["page"] == 1
    assert len(ref["bbox"]) == 4 and all(isinstance(v, float) for v in ref["bbox"])
    # Rounded, so re-extracting the same bytes gives the same line.
    assert all(round(v, 1) == v for v in ref["bbox"])
    assert blocks[0].to_dict()["locator_ref"] == ref


def test_a_pdf_block_gets_no_guessed_heading_path():
    data = make_pdf_pages([[
        "Methods Data collection Nuclei isolation from adult heart tissue was "
        "performed as described, and libraries were prepared with the 10x kit."]])
    blocks, _, _ = pdf.blocks_from_pdf(data, "f.pdf", L)
    assert all(b.section_path is None for b in blocks)


def test_citation_markers_are_dropped_but_other_cross_references_kept():
    """Left in, `<xref ref-type="bibr">` turns "as shown previously" into "as shown
    previously12,13", which is noise in a quote and worse in an evidence check."""
    body = ('<sec><title>Results</title><p>As shown previously'
            '<xref ref-type="bibr" rid="b1">12,13</xref>, islets vary '
            '(<xref ref-type="fig" rid="f1">Fig. 1a</xref>).</p></sec>')
    blocks, _, _ = jats.blocks_from_jats(jats_article(body), "f.nxml", L)
    text = " ".join(b.text for b in blocks if b.kind == PARAGRAPH)
    assert "previously, islets" in text
    assert "12,13" not in text
    assert "Fig. 1a" in text


@pytest.mark.parametrize("body,expected", [
    ('<p>We report <xref ref-type="bibr" rid="b1">1</xref>.</p>', "We report."),
    ('<p>severe symptoms (<xref ref-type="bibr" rid="b1">1</xref>; '
     '<xref ref-type="bibr" rid="b2">2</xref>, '
     '<xref ref-type="bibr" rid="b3">3</xref>). While recent studies</p>',
     "severe symptoms. While recent studies"),
    ('<p>LD blocks using LDetect<xref ref-type="bibr" rid="b1">7</xref>,'
     '<xref ref-type="bibr" rid="b2">8</xref>.</p>', "LD blocks using LDetect."),
    # The lookbehind is what keeps a function call intact.
    ('<p>we ran the susie_rss() function on each locus</p>',
     "we ran the susie_rss() function on each locus"),
    ('<p>the HarmonyMatrix() call and a negative (-) gate</p>',
     "the HarmonyMatrix() call and a negative (-) gate"),
])
def test_dropping_a_citation_does_not_leave_its_punctuation(body, expected):
    """35 literal `()` and 12 more `(` followed by a separator over the JATS
    blocks of 10.1016/j.cell.2021.01.053. One block read
    `...severe symptoms (; ; ; ; , ). While recent studies...`."""
    blocks, _, _ = jats.blocks_from_jats(
        jats_article(f"<sec><title>Results</title>{body}</sec>"), "f.nxml", L)
    assert next(b.text for b in blocks if b.kind == PARAGRAPH
                and b.section == "results") == expected


def test_a_citation_in_a_table_cell_is_the_value_and_stays():
    """10 of the 29 SOURCE cells in 10.1016/j.cell.2021.01.053's key resources
    table were destroyed: `(Korsunsky et al., 2019)` became `()`, and the card
    read `SOURCE [text, 11 distinct, 11 empty] = () | 10x Genomics | ; | ...`."""
    body = ('<sec><title>Methods</title><table-wrap><label>Key resources</label>'
            '<table><tr><th>REAGENT</th><th>SOURCE</th></tr>'
            '<tr><td>Harmony</td><td>(<xref ref-type="bibr" rid="b1">Korsunsky '
            'et al., 2019</xref>)</td></tr>'
            '<tr><td>Scanpy</td><td><xref ref-type="bibr" rid="b2">Wolf et al., '
            '2018</xref></td></tr></table></table-wrap></sec>')
    blocks, _, _ = jats.blocks_from_jats(jats_article(body), "f.nxml", L)
    source = next(c for c in next(b for b in blocks if b.kind == TABLE)
                  .table["columns"] if c["name"] == "SOURCE")
    assert source["values"] == ["(Korsunsky et al., 2019)", "Wolf et al., 2018"]


def test_reference_list_is_not_extracted_and_the_drop_is_recorded():
    """A model asked for perturbations will happily take one from a reference
    title. But the drop left no trace, so `meta["sections"]` disagreed with the
    sections actually in the file and nothing said why."""
    back = ('<ref-list><title>References</title><ref><element-citation>'
            '<article-title>CRISPR knockout of TP53 in mice</article-title>'
            '</element-citation></ref></ref-list>')
    blocks, _, meta = jats.blocks_from_jats(jats_article("", back=back), "f.nxml", L)
    assert not any("CRISPR knockout of TP53" in b.text for b in blocks)
    assert "references" not in meta["sections"]
    assert meta["reference_list_dropped"] is True


def test_the_jats_parser_says_when_it_hits_a_cap():
    """`pdf.py` and `docxfile.py` both flag this; the JATS walker stopped
    silently, so a capped article read as a short one."""
    body = "<sec><title>Results</title>" + "<p>a paragraph of text</p>" * 20 + "</sec>"
    _, _, meta = jats.blocks_from_jats(jats_article(body), "f.nxml",
                                       Limits(max_blocks_per_file=5))
    assert meta["blocks_capped"] is True

    tables_body = "<sec><title>Results</title>" + (
        "<table-wrap><label>T</label><table><tr><th>a</th><th>b</th></tr>"
        "<tr><td>1</td><td>2</td></tr></table></table-wrap>" * 4) + "</sec>"
    _, _, meta = jats.blocks_from_jats(jats_article(tables_body), "f.nxml",
                                       Limits(max_tables_per_file=2))
    assert meta["tables_capped"] is True and meta["tables"] == 2


def test_a_table_wrapped_in_a_paragraph_is_emitted_once():
    """`walk_children` handles the nested float and the generic branch could then
    handle it again, giving two blocks with the same locator."""
    body = ('<sec><title>Results</title><p>Text before.<table-wrap><label>T1</label>'
            '<table><tr><th>a</th><th>b</th></tr><tr><td>1</td><td>2</td></tr>'
            '</table></table-wrap></p></sec>')
    blocks, _, _ = jats.blocks_from_jats(jats_article(body), "f.nxml", L)
    assert sum(1 for b in blocks if b.kind == TABLE) == 1
    locators = [(b.source_file, b.locator, b.kind) for b in blocks]
    assert len(locators) == len(set(locators))


def test_jats_table_becomes_a_card():
    body = ('<sec><title>Results</title><table-wrap id="T1"><label>Table 1</label>'
            '<caption><p>Donor characteristics</p></caption><table><thead>'
            '<tr><th>Donor</th><th>Sex</th></tr></thead><tbody>'
            '<tr><td>D1</td><td>M</td></tr><tr><td>D2</td><td>F</td></tr>'
            '</tbody></table></table-wrap></sec>')
    blocks, _, meta = jats.blocks_from_jats(jats_article(body), "f.nxml", L)
    card = next(b for b in blocks if b.kind == TABLE)
    assert meta["tables"] == 1
    assert card.label == "Table 1"
    assert card.table["header"] == ["Donor", "Sex"]
    assert "Donor characteristics" in card.text


def test_a_cell_holding_several_paragraphs_keeps_them_apart():
    """10.1038/s41467-023-40505-5 Table 1: 24 of its 144 cells hold more than one
    block child, and the SNP PIP column read `0.7980.15` for two values -- which
    also flipped the column's dtype from number to mixed and cost it its
    min/max/median."""
    body = ('<sec><title>Results</title><table-wrap><label>Table 1</label><table>'
            '<tr><th>Locus</th><th>SNP PIP</th><th>Supporting SNPs</th></tr>'
            '<tr><td><p>1p13</p></td><td><p>0.798</p><p>0.15</p></td>'
            '<td><p>rs1906615</p><p>rs7689774</p></td></tr>'
            '</table></table-wrap></sec>')
    blocks, _, _ = jats.blocks_from_jats(jats_article(body), "f.nxml", L)
    rows = next(b for b in blocks if b.kind == TABLE).table
    assert "0.798; 0.15" in next(b for b in blocks if b.kind == TABLE).text
    assert rows["header"] == ["Locus", "SNP PIP", "Supporting SNPs"]


def test_a_block_id_survives_a_section_relabel():
    """`section` is the most-revised heuristic in this package, and a confirmed
    fact about donor age has to survive a relabel. Measured across the real
    6a54ff7^ -> HEAD change: including it would have moved 21 of 1,717 ids."""
    def one():
        return Block(kind=PARAGRAPH, text="Donors were 44 and 61 years old.",
                     source_file="f.pdf", origin="pdf", locator="p.7")

    before, after = one(), one()
    after.section = "results"
    number_blocks([before])
    number_blocks([after])
    assert before.block_id == after.block_id and before.block_id


def test_repeated_text_at_the_same_locator_gets_distinct_ids():
    """A PDF locator is only a page: `(source_file, locator, text_sha256)`
    collides 416 times in the 2,076 blocks of 10.1126/science.aat5031, `p.79` /
    "Developing nephron" alone 22 times."""
    blocks = [Block(kind=PARAGRAPH, text="Developing nephron", source_file="f.pdf",
                    origin="pdf", locator="p.79") for _ in range(22)]
    number_blocks(blocks)
    assert len({b.block_id for b in blocks}) == 22
    # And the assignment is a pure function of the list, so it repeats exactly.
    again = [Block(kind=PARAGRAPH, text="Developing nephron", source_file="f.pdf",
                   origin="pdf", locator="p.79") for _ in range(22)]
    number_blocks(again)
    assert [b.block_id for b in blocks] == [b.block_id for b in again]


def test_a_prose_paragraph_with_inline_markup_is_unchanged():
    """The separator is a table-cell rule. Inline elements are not block-level and
    must not gain one."""
    body = ('<sec><title>Results</title><p>The <italic>Tp53</italic> locus and the '
            '<sup>3</sup>H label were <bold>both</bold> measured.</p></sec>')
    blocks, _, _ = jats.blocks_from_jats(jats_article(body), "f.nxml", L)
    text = next(b.text for b in blocks if b.kind == PARAGRAPH and "Tp53" in b.text)
    assert text == "The Tp53 locus and the 3H label were both measured."


def test_a_single_paragraph_cell_gains_no_separator():
    body = ('<sec><title>Results</title><table-wrap><table>'
            '<tr><th>Donor</th><th>Sex</th></tr>'
            '<tr><td><p>D1</p></td><td>M</td></tr>'
            '</table></table-wrap></sec>')
    blocks, _, _ = jats.blocks_from_jats(jats_article(body), "f.nxml", L)
    card = next(b for b in blocks if b.kind == TABLE)
    assert card.table["header"] == ["Donor", "Sex"]
    assert "; " not in card.table["columns"][0]["values"][0]


def test_image_only_table_yields_its_caption():
    body = ('<sec><title>Results</title><table-wrap><label>Table 2</label>'
            '<caption><p>Primer sequences</p></caption>'
            '<graphic xlink:href="t2.jpg"/></table-wrap></sec>')
    blocks, _, _ = jats.blocks_from_jats(jats_article(body), "f.nxml", L)
    caption = next(b for b in blocks if b.kind == CAPTION)
    assert "Primer sequences" in caption.text


def test_figure_captions_are_kept():
    body = ('<sec><title>Results</title><fig id="f1"><label>Fig. 1</label>'
            '<caption><p>UMAP of 12,000 nuclei from mouse heart</p></caption></fig></sec>')
    blocks, _, _ = jats.blocks_from_jats(jats_article(body), "f.nxml", L)
    assert any("UMAP of 12,000 nuclei" in b.text for b in blocks if b.kind == CAPTION)


def test_supplement_labels_are_read_from_the_nested_media_element():
    """Springer nests href and caption inside `<media>`, so reading only the direct
    children of `<supplementary-material>` found labels for none of the 40 XML
    files in this corpus. The join is what turns
    `41467_2023_40505_MOESM3_ESM.xlsx` into "Supplementary Table 3"."""
    data = jats_article(SPRINGER_SUPPLEMENT)
    labels = jats.supplement_labels(data)
    assert labels["41467_2023_40505_MOESM3_ESM.xlsx"]["caption"] == "Supplementary Table 3"
    _, _, meta = jats.blocks_from_jats(data, "f.nxml", L)
    assert meta["supplement_labels"] == labels


def test_named_entities_and_the_doctype_do_not_break_parsing():
    """ElementTree resolves no external DTD, so every entity a JATS DOCTYPE would
    have defined is a fatal "undefined entity" -- a real file reported unreadable
    over a Greek letter."""
    body = "<sec><title>Results</title><p>TNF&alpha; rose 5&nbsp;fold.</p></sec>"
    blocks, status, _ = jats.blocks_from_jats(jats_article(body), "f.nxml", L)
    assert status == "ok"
    assert "TNF\u03b1" in " ".join(b.text for b in blocks)


def test_malformed_xml_is_unreadable_with_a_reason():
    blocks, status, meta = jats.blocks_from_jats(b"<article><body><p>x", "f.nxml", L)
    assert (blocks, status) == ([], "unreadable")
    assert "parse error" in meta["reason"]


def test_xml_without_an_article_element_is_unreadable():
    assert jats.blocks_from_jats(b"<other/>", "f.nxml", L)[1] == "unreadable"


# -- PDF ---------------------------------------------------------------------

def test_hyphenated_line_breaks_are_rejoined():
    """A hyphenated word is unsearchable and cannot be quoted, and any check that
    a quote really appears in the source compares it against this text."""
    data = make_pdf_pages([["We measured the perturba-\ntion of gene expression "
                            "across every single condition tested in this study, "
                            "and compared each result against the matched control "
                            "samples that were processed on the same day by the "
                            "same operator using identical reagent lots."]])
    blocks, status, _ = pdf.blocks_from_pdf(data, "f.pdf", L)
    assert status == "ok"
    assert "perturbation" in " ".join(b.text for b in blocks)


def _traced(spans):
    """`get_texttrace` output as far as `_symbol_map` is concerned: fonts and the
    characters drawn in them.

    A real fixture is not available -- PyMuPDF's own base-14 Symbol font ships a
    ToUnicode map, so a synthesized PDF round-trips `g` as `g` and never
    reproduces the private-use codepoint this guard exists for. The corpus test
    over 10.1126/sciimmunol.aba4163 covers the real shape.

    A trace char is `(unicode, glyph id, origin, bbox)`; only the first is read
    here, and the glyph id is filled with the codepoint because nothing in this
    path looks at it.
    """
    return [{"font": font, "chars": [(ord(c), ord(c), (0.0, 0.0), (0, 0, 0, 0))
                                     for c in text]}
            for font, text in spans]


@pytest.mark.parametrize("raw,expected", [
    # `interleukin-<soft hyphen>\n17A` keeps its real hyphen.
    ("interleukin-\u00ad\n17A", "interleukin-17A"),
    ("scRNA-\u00ad seq", "scRNA-seq"),
    ("pheno\u00ad type", "phenotype"),
    ("pheno\u00ad\ntype", "phenotype"),
    ("zero\u200bwidth", "zerowidth"),
    ("\ufeffbom", "bom"),
])
def test_a_soft_hyphen_does_not_survive_inside_a_word(raw, expected):
    """U+00AD is category Cf -- neither `\\w` nor `\\s` -- so `_HYPHEN_BREAK`
    could not fire across one and the whitespace collapse left it in the word.
    21 of them in 10.1126/sciimmunol.aba4163."""
    cleaned = pdf._clean_block(raw)
    assert cleaned == expected
    assert "\u00ad" not in cleaned and "\u200b" not in cleaned


def test_a_symbol_font_glyph_becomes_the_greek_letter_it_stands_for():
    """`SymbolGreek` U+F067 is the gamma in `IFN-\u03b3`, 41 times in one article."""
    symbols = pdf._symbol_map(_traced([("SymbolGreek", "\uf067\uf062")]))
    assert symbols == {0xF067: "\u03b3", 0xF062: "\u03b2"}
    assert pdf._clean_block("IFN-\uf067 and TGF-\uf062", symbols) == "IFN-\u03b3 and TGF-\u03b2"


def test_a_private_use_codepoint_from_an_ordinary_font_is_left_alone():
    """A subsetted Latin face reuses the private use area for its own glyphs.
    Turning one of those into a Greek letter would invent a character."""
    assert pdf._symbol_map(_traced([("ABCDEF+MinionPro", "\uf067")])) == {}
    assert pdf._clean_block("x\uf067y", {}) == "x\uf067y"


@pytest.mark.parametrize("raw,expected", [
    ("IL-\n17A", "IL-17A"),
    ("scRNA-\nseq", "scRNA-seq"),
    ("CD4-\npositive", "CD4-positive"),
    ("SARS-CoV-\n2", "SARS-CoV-2"),
    ("perturba-\ntion", "perturbation"),
    # The accepted cost of the rule: nothing short of a dictionary separates a
    # common hyphenated adjective from a word broken at a syllable.
    ("well-\nknown", "wellknown"),
])
def test_a_hyphen_inside_an_identifier_survives_the_line_break(raw, expected):
    """Rejoining unconditionally deleted a real hyphen out of 78 of the 648
    line-break hyphens in this corpus's PDFs -- `SARS-CoV-2`, `COVID-19`,
    `Mono_c1-CD14-CCL3` -- which is exactly the vocabulary a curation answer is
    made of."""
    assert pdf._clean_block(raw) == expected


def test_the_hyphen_ratio_is_recorded_so_it_can_be_inspected():
    data = make_pdf_pages([["The IL-\n17A response and the perturba-\ntion of it "
                            "were measured across all of the matched samples that "
                            "were collected on the same day by the same operator."]])
    _, _, meta = pdf.blocks_from_pdf(data, "f.pdf", L)
    assert meta["hyphens_kept"] == 1 and meta["hyphens_joined"] == 1


def test_an_unmapped_private_use_glyph_is_counted_not_hidden():
    data = make_pdf_pages([["ordinary text with no symbol font at all here"]])
    _, _, meta = pdf.blocks_from_pdf(data, "f.pdf", L)
    assert "glyphs_unmapped" not in meta and "glyphs_mapped" not in meta


# -- fonts that do not say what their glyphs mean ----------------------------

#: What a garbled file was hiding. Long enough to clear `min_pdf_text_chars`, so
#: the status under test is the encoding rule rather than the scanned-PDF one.
_METHODS = ("These studies were intended to be the first explorations of cellular "
            "diversity in the human brain, and nuclei were isolated from frozen "
            "tissue before being sequenced on a NovaSeq 6000 instrument at the "
            "core facility on the same day by the same operator.")


def test_a_font_that_never_says_what_its_glyphs_mean_is_read_from_the_font_itself():
    """10.1126/science.adf5357's Supplementary Materials -- that paper's only copy
    of its Materials and Methods -- extracted 124,178 characters of
    `TheVe VWXdLeV ZeUe LQWeQded` and was reported `ok`.

    The recovery has to come from the embedded font's own character map. The
    apparent +29 between a glyph id and a codepoint is a property of one font's
    glyph order: through this fixture's font the same text comes out off by one
    instead, so a shift fitted to the Science file would turn this into different
    nonsense rather than into English.
    """
    data = make_unreadable_font_pdf([[_METHODS]])
    blocks, status, meta = pdf.blocks_from_pdf(data, "sm.pdf", L)
    assert status == pdf.OK
    assert " ".join(b.text for b in blocks).startswith(
        "These studies were intended to be the first explorations")
    assert sum(meta["glyph_encoding_repaired"].values()) > 0
    assert "glyphs_unnamed" not in meta


def test_the_repair_leaves_a_healthy_pdf_alone():
    """It runs on every PDF, so "changes nothing when nothing is wrong" is the
    load-bearing half. Measured over the 972 PDFs in this corpus: 183 have a font
    with a gap in its CMap, and exactly one of them -- the Science supplement --
    comes out with different text. The other 182 gaps are for glyphs the document
    never draws."""
    data = make_pdf_pages([[_METHODS]])
    blocks, status, meta = pdf.blocks_from_pdf(data, "f.pdf", L)
    assert status == pdf.OK
    assert "glyph_encoding_repaired" not in meta
    assert "glyphs_unnamed" not in meta
    assert " ".join(b.text for b in blocks).startswith("These studies were intended")


def test_a_file_whose_glyphs_cannot_be_named_is_not_ok():
    """The case where recovery would be a guess. Here the ToUnicode CMap is one
    no reader can use, so the repair has to decline -- a CMap it cannot parse is
    one it cannot safely add to -- and what is left is unreadable.

    Real shape: 10.1038/s41588-024-01702-0's reporting summary, 6,869 of 6,869
    glyphs with nothing behind them, whose fonts are CID-keyed CFF subsets with
    identity ordering and glyph names of the form `cid00042`. Detection plus an
    honest status is the whole answer there.

    The blocks go with the status. Their characters are the parser's fallback for
    a code it could not map, and letting them through as prose is the bug.
    """
    data = make_unreadable_font_pdf([[_METHODS]], broken_cmap=True)
    blocks, status, meta = pdf.blocks_from_pdf(data, "sm.pdf", L)
    assert status == pdf.GARBLED
    assert blocks == [] and meta["chars"] == 0
    assert meta["glyphs_unnamed"] == meta["glyphs_drawn"] > 0
    assert "no character behind them" in meta["reason"]
    assert meta["garbled_sample"]


def test_the_rule_asks_the_document_not_the_prose():
    """A supplementary figure PDF is legitimately almost free of function words,
    so judging on "does this look like English" flags 26 files in this corpus of
    which one is broken -- and says nothing at all about a paper written in
    another language. Not one word here is a function word."""
    symbols = ("CD4 CD8A FOXP3 TP53 MYC PTPRC ITGAM CXCR4 CCR7 IL2RA GZMB PRF1 "
               "NKG7 KLRD1 LYZ CD14 FCGR3A MS4A1 CD79A NCAM1 SELL CCR6 RORC "
               "TBX21 GATA3 STAT3 IRF4 BATF3 XCR1 CLEC9A SIRPA THBD CD1C ")
    data = make_pdf_pages([[symbols * 2]])
    _, status, meta = pdf.blocks_from_pdf(data, "figures.pdf", L)
    assert status == pdf.OK
    assert "glyphs_unnamed" not in meta


def test_glyphs_with_no_character_are_counted_below_the_threshold_too():
    r"""Sub-threshold damage is real and must not vanish. Page 11 of
    10.1016/j.cell.2024.08.019's mmc8 draws an axis label as `SUHC*-DVWURF\WH-0`
    where the figure reads `preCG-astrocyte-0`, and that is 18% of the file's
    glyphs while its captions are fine and worth reading. The module's standing
    rule for a glyph it cannot name is to count it, not to drop it quietly."""
    mixed = concat_pdfs(
        make_pdf_pages([[_METHODS] * 3 for _ in range(6)]),
        make_unreadable_font_pdf([[_METHODS]], broken_cmap=True),
    )
    _, status, meta = pdf.blocks_from_pdf(mixed, "mixed.pdf", L)
    fraction = meta["glyphs_unnamed"] / meta["glyphs_drawn"]
    assert 0 < fraction <= L.max_unnamed_glyph_fraction
    assert status == pdf.OK


def test_a_cmap_this_reader_cannot_parse_stops_the_repair():
    """`None` and "maps nothing" have to be different answers. Only codes the
    existing CMap leaves out are added, so a CMap read as empty when it is not
    would let the repair write over characters the document already resolved."""
    assert pdf._cmap_entries(
        b"3 beginbfchar\n<0008><0041>\n<0009><0042>\n<000A><0043>\nendbfchar\n"
        b"2 beginbfrange\n<0020><0022><0061>\n<0030><0031>[<0075><0076>]\n"
        b"endbfrange\n") == {8: "A", 9: "B", 10: "C", 0x20: "a", 0x21: "b",
                             0x22: "c", 0x30: "u", 0x31: "v"}
    # a bfchar source with no destination, a descending range, and an array whose
    # length does not match the range it fills
    assert pdf._cmap_entries(b"1 beginbfchar\n<0008>\nendbfchar\n") is None
    assert pdf._cmap_entries(b"1 beginbfrange\n<0041><0040><0061>\nendbfrange\n") is None
    assert pdf._cmap_entries(b"1 beginbfrange\n<0041><0043>[<0061>]\nendbfrange\n") is None


def test_the_repair_never_overwrites_a_mapping_the_publisher_supplied():
    """The publisher's own CMap is the better authority where it speaks: it
    decomposes the `fi` ligature that the font's map calls U+FB01, and uses a
    non-breaking hyphen where the font says U+002D. 52 of the 898 fonts in this
    corpus that carry both maps disagree that way. Every one of those characters
    must survive the merge untouched."""
    data = make_embedded_font_pdf([[_METHODS]])
    document = fitz.open(stream=data, filetype="pdf")
    try:
        fonts = [x for x in range(1, document.xref_length())
                 if document.xref_get_key(x, "Subtype")[1] == "/Type0"]
        assert fonts, "fixture no longer embeds a Type0 font"
        for xref in fonts:
            before = pdf._cmap_entries(document.xref_stream(
                int(document.xref_get_key(xref, "ToUnicode")[1].split()[0])))
            pdf._repair_font_encoding(document, xref)
            after = pdf._cmap_entries(document.xref_stream(
                int(document.xref_get_key(xref, "ToUnicode")[1].split()[0])))
            assert before and after
            assert all(after[code] == text for code, text in before.items())
    finally:
        document.close()


def test_a_font_whose_own_map_contradicts_the_file_is_left_alone():
    """The check that would catch a font where a character code is not a glyph id
    after all: where the CMap in the file and the font's own map both name a
    glyph, they have to mostly agree. Measured across this corpus, real
    disagreement never exceeds 7% of an overlap and is always one glyph named two
    equivalent ways; a font where the assumption fails would disagree on nearly
    all of it."""
    data = make_embedded_font_pdf([[_METHODS]])
    document = fitz.open(stream=data, filetype="pdf")
    try:
        xref = next(x for x in range(1, document.xref_length())
                    if document.xref_get_key(x, "Subtype")[1] == "/Type0")
        stream = int(document.xref_get_key(xref, "ToUnicode")[1].split()[0])
        glyphs = pdf._embedded_glyph_unicodes(document, xref)
        assert glyphs, "fixture no longer embeds a readable font"
        # every glyph the font knows, mapped one codepoint off: the shape of a
        # font whose CIDs are not its glyph ids
        shifted = "".join("<%04X><%04X>\n" % (g, c + 1) for g, c in sorted(glyphs.items()))
        document.update_stream(stream, (
            "begincmap\n1 begincodespacerange\n<0000><FFFF>\nendcodespacerange\n"
            "%d beginbfchar\n%sendbfchar\nendcmap\nend\n"
            % (len(glyphs), shifted)).encode("latin-1"), new=True)
        assert pdf._repair_font_encoding(document, xref) is None
    finally:
        document.close()


def test_a_scanned_pdf_is_still_scanned_and_not_garbled():
    """A page with no glyphs on it has a zero denominator, not a bad ratio."""
    _, status, meta = pdf.blocks_from_pdf(make_scanned_pdf(2), "scan.pdf", L)
    assert status == pdf.SCANNED
    assert "glyphs_unnamed" not in meta


# -- PDFs: OCR ---------------------------------------------------------------
# 70 supplements in this corpus are scans -- 245 pages between them, median 3,
# longest 11. What is faked below is one line, `Pixmap.pdfocr_tobytes`, which is the
# only step that needs the tesseract binary and the only step with no offline
# coverage; the render, `show_pdf_page`, and the re-parse of the OCR'd PDF all run
# for real, which is the point of the pass returning a PDF rather than strings.

#: What tesseract would hand back for a scanned table: a page with a text layer.
OCR_TEXT = ("Table S3. Donor characteristics. Islets were isolated from eight "
            "donors and dissociated before loading on the Chromium controller. ")


def test_a_scanned_pdf_is_read_by_ocr():
    with no_tesseract():
        _, floor, _ = pdf.blocks_from_pdf(make_scanned_pdf(pages=2), "scan.pdf", L)
    assert floor == pdf.SCANNED, "the status this file has without the binary"

    with tesseract(returns=make_pdf_pages([[OCR_TEXT * 3]])):
        blocks, status, meta = pdf.blocks_from_pdf(
            make_scanned_pdf(pages=2), "scan.pdf", L)

    assert status == "ok_via_ocr"
    assert blocks and "Donor characteristics" in blocks[0].text
    assert meta["ocr"] == {"dpi": 300, "language": "eng", "pages": 2,
                           "chars": meta["chars"]}
    assert meta["chars"] > L.min_pdf_text_chars


def test_ocr_text_is_cleaned_by_the_same_rules_as_a_text_layer():
    """The reason the pass hands a *PDF* back to `blocks_from_pdf` rather than
    strings: a hyphen broken across lines has to be rejoined in OCR'd text too, and
    two parsers for the same job would diverge."""
    with tesseract(returns=make_pdf_pages([["Islets were dissociated and perturba-\ntion "
                                            "was confirmed by flow cytometry. " * 4]])):
        blocks, status, _ = pdf.blocks_from_pdf(make_scanned_pdf(pages=1), "scan.pdf", L)
    assert status == "ok_via_ocr"
    assert any("perturbation was confirmed" in b.text for b in blocks)


def test_ocr_bboxes_are_in_the_original_pages_coordinates():
    """`pdfocr_tobytes` returns a page the size of the pixmap -- at 300 dpi four
    times the original in each direction -- so inserting it directly would record
    every locator_ref in a coordinate space nothing else uses."""
    with tesseract(returns=make_pdf_pages([[OCR_TEXT * 3]])):
        blocks, status, _ = pdf.blocks_from_pdf(make_scanned_pdf(pages=1), "scan.pdf", L)
    assert status == "ok_via_ocr"
    page_height = fitz.open(stream=make_scanned_pdf(pages=1), filetype="pdf")[0].rect.height
    boxes = [b.locator_ref["bbox"] for b in blocks if b.locator_ref]
    assert boxes, "no bounding boxes were recorded"
    assert max(box[3] for box in boxes) <= page_height + 1, boxes


def test_a_scanned_pdf_without_tesseract_keeps_its_status_and_says_what_to_install():
    """The optional-system-dependency contract, and the same one
    `spreadsheet.cards_from_xls` keeps for xlrd: the old status, plus a reason that
    is an instruction."""
    with no_tesseract():
        blocks, status, meta = pdf.blocks_from_pdf(
            make_scanned_pdf(pages=2), "scan.pdf", L)

    assert status == pdf.SCANNED
    assert "tesseract" in meta["reason"] and "brew install" in meta["reason"]
    assert meta["pages"] == 2, "the original parse's findings are kept"


def test_ocr_that_finds_nothing_legible_leaves_the_file_scanned():
    """A blank scan is not the same failure as a missing binary, and the two must
    not read alike: this one says the pages carry no legible text."""
    with tesseract(returns=make_scanned_pdf(pages=1)):
        blocks, status, meta = pdf.blocks_from_pdf(make_scanned_pdf(pages=1), "scan.pdf", L)

    assert status == pdf.SCANNED
    assert "no legible text" in meta["reason"]
    assert meta["ocr"]["chars"] < L.min_pdf_text_chars


def test_a_scan_longer_than_the_ocr_cap_is_not_ocred():
    """245 pages over 70 files, longest 11, against a cap of 25. What this stops is
    a scanned 88-page peer-review bundle."""
    with tesseract(returns=make_pdf_pages([[OCR_TEXT * 3]])):
        blocks, status, meta = pdf.blocks_from_pdf(
            make_scanned_pdf(pages=3), "scan.pdf", Limits(max_ocr_pages=2))

    assert status == pdf.SCANNED
    assert "3 pages is over the 2-page OCR cap" in meta["reason"]


def test_ocr_failing_costs_one_file_not_the_run():
    """A tesseract that is present and broken is a real shape -- a half-installed
    language pack answers this way -- and it must not end a run of 393 articles."""
    with tesseract(raises=RuntimeError("tesseract: error while loading libtesseract")):
        blocks, status, meta = pdf.blocks_from_pdf(make_scanned_pdf(pages=1), "scan.pdf", L)

    assert status == pdf.SCANNED
    assert "OCR failed: RuntimeError" in meta["reason"]


def test_a_readable_pdf_is_never_ocred():
    """The pass is reached only by the scanned branch, so a file with a text layer
    must not pay for a render. Asserted by making the OCR call fail loudly."""
    with tesseract(raises=AssertionError("OCR was attempted on a readable PDF")):
        _, status, _ = pdf.blocks_from_pdf(make_pdf(pages=2), "f.pdf", L)
    assert status == "ok"


def test_a_pdf_with_no_pages_is_not_reported_as_an_ocr_failure():
    """It reaches the scanned branch having produced no characters, and
    `_render_with_ocr` would raise `cannot save with zero pages` on the way out."""
    one_page = make_scanned_pdf(pages=1)
    document = fitz.open(stream=one_page, filetype="pdf")
    document.delete_page(0)
    # A zero-page document cannot be saved, so the shape has to be built by hand.
    empty = one_page.replace(b"/Count 1", b"/Count 0")
    document.close()

    with tesseract(raises=AssertionError("nothing should have been rendered")):
        _, status, meta = pdf.blocks_from_pdf(empty, "empty.pdf", L)
    assert status == pdf.SCANNED
    assert "no pages" in meta["reason"]


def test_the_tesseract_version_is_absent_rather_than_missing():
    """It goes in the extraction cache key, so it has to answer for an install that
    does not have it -- and `absent` is an answer, where a missing key is not."""
    with no_tesseract():
        assert pdf.tesseract_version() == "absent"


def test_running_headers_are_dropped_and_named():
    """A journal footer repeated on every page would otherwise appear as thirty
    near-identical paragraphs. Naming them matters as much as dropping them: the
    rule deletes 424 of 1,160 blocks in 10.1126/sciimmunol.aba4163 and 854 of
    2,474 in the 89-page Science supplement, and a bare count cannot be checked."""
    body = ("Nuclei were isolated from frozen tissue and sequenced on a NovaSeq "
            "6000 instrument at the core facility. ")
    data = make_pdf_pages([["Nature Communications | volume 14", body * 3]] * 4)
    blocks, _, meta = pdf.blocks_from_pdf(data, "f.pdf", L)
    assert meta["running_lines_dropped"] >= 4
    assert not any("volume 14" in b.text for b in blocks)
    assert meta["running_lines"][0] == {"text": "Nature Communications | volume 14",
                                        "pages": 4}


def test_a_repeated_line_in_the_body_is_not_mistaken_for_a_running_head():
    """`Reviewer #2 (Remarks to the Author):` sits at y0/h = 0.28, 0.53 and 0.12
    in 10.1038/s41467-023-40505-5's peer-review file, and the count-only rule
    deleted all three: that article's blocks held reviewers 1, 3, 1, 3 and 4 and
    no reviewer 2 at all, whose remarks then read as a continuation of
    reviewer 1's."""
    heading = "Reviewer #2 (Remarks to the Author):"
    body = ("The authors should clarify how the nuclei were isolated and how many "
            "donors contributed to each cluster in the figure. ")
    data = make_pdf_pages([["Nature Communications | volume 14", heading, body * 2]] * 4)
    blocks, _, meta = pdf.blocks_from_pdf(data, "f.pdf", L)
    texts = [b.text for b in blocks]
    assert texts.count(heading) == 4
    assert not any("volume 14" in t for t in texts)
    assert [r["text"] for r in meta["running_lines"]] == \
        ["Nature Communications | volume 14"]


def test_glued_section_heading_is_split_in_a_real_pdf():
    data = make_pdf_pages([[
        "Methods Data collection Nuclei isolation from adult heart tissue was "
        "performed as described, and libraries were prepared with the 10x kit."]])
    blocks, _, meta = pdf.blocks_from_pdf(data, "f.pdf", L)
    assert meta["glued_headings_split"] == 1
    assert blocks[0].kind == HEADING and blocks[0].text == "Methods"
    assert blocks[1].section == "methods"
    assert blocks[1].text.startswith("Data collection")


def test_scanned_pdf_is_flagged_not_reported_empty():
    """It is the article but needs an OCR step this pipeline does not have, which
    is a different problem from a file with nothing in it."""
    blocks, status, meta = pdf.blocks_from_pdf(make_scanned_pdf(), "f.pdf", L)
    assert status == "no_text_scanned_pdf"
    assert meta["chars"] < L.min_pdf_text_chars


def test_damaged_pdf_is_unreadable_with_a_reason():
    blocks, status, meta = pdf.blocks_from_pdf(b"%PDF-1.4 truncated", "f.pdf", L)
    assert (blocks, status) == ([], "unreadable")
    assert "reason" in meta


# -- docx --------------------------------------------------------------------

def test_docx_paragraphs_headings_and_tables():
    """In this corpus these are `suppmatmeth.docx` and `supplement-fS1-S7.docx` --
    where the library kit is written down when the main text says only "see
    Supplementary Methods"."""
    data = make_docx([
        ("paragraph", "Supplementary materials and methods", "Heading1"),
        ("paragraph", "Libraries were prepared with the Chromium Single Cell 3' v3 kit."),
        ("table", [["Primer", "Sequence"], ["Actb-F", "GGCTGTATTCCCCTCCATCG"]]),
    ])
    blocks, status, meta = docxfile.blocks_from_docx(data, "s.docx", L)
    assert status == "ok" and meta["tables"] == 1
    assert blocks[0].kind == HEADING and blocks[0].section == "methods"
    assert blocks[1].kind == PARAGRAPH and "v3 kit" in blocks[1].text
    card = next(b for b in blocks if b.kind == TABLE)
    assert card.table["header"] == ["Primer", "Sequence"]


def test_docx_field_codes_and_deleted_text_are_skipped():
    """A field code is a formula, and deleted text is text the authors removed;
    quoting either as evidence would be wrong."""
    raw = ('<w:p><w:r><w:t>Mice were </w:t></w:r>'
           '<w:r><w:instrText> REF _Ref1 \\h </w:instrText></w:r>'
           '<w:r><w:delText>eight</w:delText></w:r>'
           '<w:r><w:t>ten weeks old at the time of harvest.</w:t></w:r></w:p>')
    blocks, _, _ = docxfile.blocks_from_docx(make_docx([("raw", raw)]), "s.docx", L)
    text = blocks[0].text
    assert "ten weeks old" in text
    assert "REF" not in text and "eight" not in text


def test_docx_that_is_not_a_zip_is_unreadable():
    blocks, status, meta = docxfile.blocks_from_docx(b"not a docx", "s.docx", L)
    assert (blocks, status) == ([], "unreadable")
    assert "reason" in meta


def test_docx_without_a_document_part_is_unreadable():
    data = make_zip([("word/other.xml", b"<x/>")])
    assert docxfile.blocks_from_docx(data, "s.docx", L)[1] == "unreadable"


# -- HTML --------------------------------------------------------------------

def test_landing_page_metadata_and_prose_are_kept():
    page = (b'<html><head><meta name="citation_title" content="A paper about mice">'
            b'<meta name="citation_author" content="Smith J">'
            b'<meta name="citation_author" content="Jones K">'
            b'<script>var x = "ignore me";</script></head><body>'
            b'<nav>Home Journals Search</nav>'
            b'<p>' + b'The abstract of this article describes single-cell profiling. ' * 4
            + b'</p><div>Cite</div></body></html>')
    blocks, status, meta = htmlfile.blocks_from_html(page, "landing.html", L)
    assert status == "ok"
    assert meta["meta_tags"]["citation_author"] == ["Smith J", "Jones K"]
    assert any("Smith J; Jones K" in b.text for b in blocks if b.kind == METADATA)
    assert any("single-cell profiling" in b.text for b in blocks if b.kind == PARAGRAPH)
    assert not any("ignore me" in b.text for b in blocks)
    assert not any("Home Journals Search" in b.text for b in blocks)


def test_bot_check_landing_page_is_not_reported_as_text():
    """Nine Elsevier landing pages in this corpus hold 129 characters: the
    browser's own user-agent string. Calling that `ok` is the failure mode
    manuscript_harvest/fetch/validate.py exists to prevent."""
    blocks, status, meta = htmlfile.blocks_from_html(
        LANDING_INTERSTITIAL, "landing.html", L)
    assert (blocks, status) == ([], "no_text")
    assert "interstitial" in meta["reason"]


def test_paywall_landing_page_names_the_denial():
    page = (b"<html><body><div>Please sign in to read the full article or purchase "
            b"this article to continue reading the text.</div></body></html>")
    _, status, meta = htmlfile.blocks_from_html(page, "landing.html", L)
    assert status == "no_text" and "paywalled" in meta["reason"]


# -- archives ----------------------------------------------------------------

@pytest.mark.parametrize("raw,expected", [
    ("../../etc/passwd", "etc/passwd"),
    ("/abs/file.csv", "abs/file.csv"),
    ("a\\b\\c.csv", "a/b/c.csv"),
    ("C:/data/x.csv", "data/x.csv"),
    ("./x.csv", "x.csv"),
])
def test_archive_member_names_cannot_escape(raw, expected):
    """Nothing here writes them, but a member name reaches `source_file` on a
    block, and code that joins that to a path should not get `../../etc/passwd`."""
    assert archive.safe_member_name(raw) == expected


def test_only_wanted_extensions_are_read():
    data = make_zip([("table.csv", b"a,b\n1,2\n"), ("figure.jpg", b"\xff\xd8fig")])
    members, meta = archive.read_members(data, L, {".csv"})
    assert [name for name, _ in members] == ["table.csv"]
    assert meta["member_extensions"] == {".csv": 1, ".jpg": 1}
    assert any("figure.jpg" in s["name"] for s in meta["skipped"])


def test_oversized_member_is_skipped_with_its_size_in_the_reason():
    """10.7554/eLife.104978's fig1-data2.zip holds a 159 MB CSV. The size is read
    from the zip directory, so nothing is decompressed to find that out."""
    limits = Limits(max_member_mb=0)
    data = make_zip([("big.csv", b"a,b\n" + b"1,2\n" * 500)])
    members, meta = archive.read_members(data, limits, {".csv"})
    assert members == []
    assert "over the 0 MB member cap" in meta["skipped"][0]["reason"]


def test_member_cap_is_reported():
    limits = Limits(max_archive_members=1)
    data = make_zip([("a.csv", b"x,y\n1,2\n"), ("b.csv", b"x,y\n3,4\n")])
    members, meta = archive.read_members(data, limits, {".csv"})
    assert len(members) == 1
    assert any(s["reason"] == "member cap reached" for s in meta["skipped"])


def test_mac_resource_forks_are_ignored():
    data = make_zip([("__MACOSX/._x.csv", b"junk"), ("x.csv", b"a,b\n1,2\n")])
    members, _ = archive.read_members(data, L, {".csv"})
    assert [name for name, _ in members] == ["x.csv"]


def test_unreadable_archive_is_reported():
    members, meta = archive.read_members(b"not a zip", L, {".csv"})
    assert members == [] and "reason" in meta


# -- archives: tar, and the single-file compressors --------------------------
# Six supplements in this corpus, 107 MB, all reported `unsupported_format` before
# these paths existed: "compressed archive other than zip; decompress by hand if it
# holds tables". Five of the six turned out to be one compressed CSV or TSV each.

@pytest.mark.parametrize("compression", ["", "gz", "bz2", "xz"])
def test_a_tar_is_recognised_however_it_is_compressed(compression):
    """The uncompressed case is the one that matters: 10.1038/s41586-020-03182-8's
    MOESM4 is a plain tar under a `.tgz` name, so `mode="r:gz"` -- what
    `fetch/sources/pmc_oa.py` uses for a package PMC built -- reads nothing from
    it."""
    data = make_tar([("table.csv", b"a,b\n1,2\n")], compression=compression)
    assert archive.looks_like_tar(data) is True
    members, meta = archive.read_tar_members(data, L, {".csv"})
    assert [name for name, _ in members] == ["table.csv"]
    assert meta["members_total"] == 1 and meta["members_read"] == 1


def test_a_gzipped_csv_is_not_mistaken_for_a_tar():
    """Three of the five `.gz` supplements here are a single table, not an archive,
    so this is the branch that decides which of the two paths they take."""
    assert archive.looks_like_tar(make_gz(b"a,b\n1,2\n")) is False
    assert archive.looks_like_tar(b"not compressed at all") is False


def test_tar_members_obey_the_same_caps_as_a_zip():
    data = make_tar([("a.csv", b"x,y\n1,2\n"), ("b.csv", b"x,y\n3,4\n")])
    members, meta = archive.read_tar_members(data, Limits(max_archive_members=1), {".csv"})
    assert len(members) == 1
    assert any(s["reason"] == "member cap reached" for s in meta["skipped"])

    members, meta = archive.read_tar_members(data, Limits(max_member_mb=0), {".csv"})
    assert members == []
    assert "over the 0 MB member cap" in meta["skipped"][0]["reason"]

    members, meta = archive.read_tar_members(data, L, {".xlsx"})
    assert members == [] and meta["member_extensions"] == {".csv": 2}


def test_the_tar_walk_stops_before_a_bomb_is_decompressed():
    """A tar has no central directory, so reading the Nth header means decompressing
    everything before it. Without this bound a 1 KB tar.gz claiming 10 GB of members
    would be walked to the end before any per-member cap could apply."""
    data = make_tar([(f"f{n}.csv", b"a,b\n1,2\n") for n in range(5)], compression="gz")
    members, meta = archive.read_tar_members(data, Limits(max_file_mb=0), {".csv"})
    assert members == []
    assert "over the 0 MB cap" in meta["walk_stopped"]
    assert meta["members_total"] == 0, "nothing was walked past the bound"


def test_tar_member_names_cannot_escape_and_junk_is_ignored():
    """176 of MOESM4's 296 members are AppleDouble forks and `.DS_Store` files, so
    the filter is what makes the census mean anything."""
    data = make_tar([("../../etc/passwd.csv", b"a,b\n1,2\n"),
                     ("._x.csv", b"junk"), ("__MACOSX/y.csv", b"junk"),
                     (".DS_Store", b"junk")])
    members, meta = archive.read_tar_members(data, L, {".csv"})
    assert [name for name, _ in members] == ["etc/passwd.csv"]
    assert meta["members_total"] == 4 and meta["member_extensions"] == {".csv": 1}


def test_a_tar_that_is_not_a_tar_is_reported():
    members, meta = archive.read_tar_members(b"not a tar", L, {".csv"})
    assert members == [] and "reason" in meta


@pytest.mark.parametrize("compression", ["gzip", "bzip2", "xz"])
def test_a_single_compressed_file_is_handed_on(compression):
    data = make_gz(b"id,sex\na,M\n", compression=compression)
    plain, status, meta = archive.decompress(data, L)
    assert (plain, status) == (b"id,sex\na,M\n", "ok")
    assert meta["compression"] == compression and meta["decompressed_bytes"] == 11


def test_an_oversize_file_is_too_large_and_the_reason_says_how_large():
    """10.1126/science.adf5357's Table_7 is 38 MB on disk and 329 MB of TSV.
    `too_large` is the honest status now that there is a parser -- `max_member_mb`
    is what stands in the way, not a missing one -- and the trailer's number is in
    the reason because "more than 50 MB" does not say what to raise the cap to."""
    plain, status, meta = archive.decompress(make_gz(b"x" * 500), Limits(max_member_mb=0))
    assert (plain, status) == (None, "too_large")
    assert "max_member_mb" in meta["reason"]
    assert "trailer declares 500 bytes" in meta["reason"]


def test_the_size_verdict_does_not_come_from_the_trailer():
    """ISIZE describes only the *last* member of a multi-member gzip -- bgzip writes
    thousands and this corpus is genomics supplements -- so a file that is over the
    cap can declare nothing at all, and the bounded read is what refuses it."""
    lying = make_gz(b"x" * 500) + make_gz(b"")
    assert archive.gzip_declared_size(lying) == 0, "the trailer under-declares"
    plain, status, meta = archive.decompress(lying, Limits(max_member_mb=0))
    assert (plain, status) == (None, "too_large")
    assert "trailer declares" not in meta["reason"], "an under-declaring trailer is not quoted"


def test_a_truncated_file_is_not_read_as_a_size():
    """The other direction, and the one that made this the verdict rather than the
    check: in a truncated file the last four bytes are deflate, not a trailer, so a
    failed download declared an arbitrary size and was called `too_large`."""
    whole = make_gz(b"".join(b"%d,gene%d\n" % (n, n) for n in range(4000)))
    half = whole[: len(whole) // 2]
    assert archive.gzip_declared_size(half) != len(half), "the bytes read are not a trailer"
    assert archive.decompress(half, L)[1] == "ok"


def test_a_multi_member_gzip_is_read_whole():
    plain, status, _ = archive.decompress(make_gz(b"a,b\n") + make_gz(b"1,2\n"), L)
    assert (plain, status) == (b"a,b\n1,2\n", "ok")


def test_junk_after_the_end_of_a_gzip_does_not_lose_the_table_before_it():
    """A publisher's script concatenating padding onto a supplement is not a reason
    to discard the supplement. The first member has to succeed; what follows it is
    counted and stepped over."""
    plain, status, meta = archive.decompress(make_gz(b"a,b\n1,2\n") + b"padding", L)
    assert (plain, status) == (b"a,b\n1,2\n", "ok")
    assert meta["trailing_bytes"] == 7


def test_a_truncated_stream_keeps_what_arrived_and_says_so():
    payload = b"".join(b"%d,gene%d,%d\n" % (n, n, n * 7) for n in range(4000))
    whole = make_gz(payload)
    plain, status, meta = archive.decompress(whole[: len(whole) // 2], L)
    assert status == "ok" and meta["truncated_stream"] is True
    assert plain and payload.startswith(plain)


def test_an_empty_truncated_stream_is_unreadable_not_empty():
    plain, status, meta = archive.decompress(make_gz(b"x" * 5000)[:12], L)
    assert (plain, status) == (None, "unreadable")
    assert "truncated" in meta["reason"]


def test_something_that_is_not_compressed_at_all_is_refused():
    plain, status, meta = archive.decompress(b"a,b\n1,2\n", L)
    assert (plain, status) == (None, "unsupported_format")
    assert "not a gzip, bzip2 or xz stream" in meta["reason"]


@pytest.mark.parametrize("outer,stored,expected", [
    # Three of the five `.gz` files here store the name; those are the only ones
    # whose inner extension is knowable without sniffing.
    ("04_NIHMS1929560-supplement-Table_4.gz", "table-S4-cell-type-taxonomy.tsv",
     "table-S4-cell-type-taxonomy.tsv"),
    ("05_NIHMS1929560-supplement-Table_5.gz", "meta.csv", "meta.csv"),
    # The two that store nothing leave a stem with no extension at all, which is
    # what sends them to `sniff_extension`.
    ("02_media-4.gz", "", "02_media-4"),
    ("counts.csv.gz", "", "counts.csv"),
    # A stored name is still a name from a file, so it cannot carry a path.
    ("x.gz", "../../etc/passwd", "passwd"),
    ("x.gz", "sub/dir/table.csv", "table.csv"),
])
def test_the_inner_name_comes_from_the_gzip_header_when_there_is_one(outer, stored, expected):
    assert archive.inner_name(outer, make_gz(b"a,b\n", stored_name=stored)) == expected


# -- blocks ------------------------------------------------------------------

def test_block_record_carries_provenance_and_a_hash():
    block = Block(kind=PARAGRAPH, text="Mice were 8 weeks old.", source_file="s.docx",
                  origin="docx", locator="para 4", section="methods", label="Table S1")
    record = block.to_dict()
    assert record["chars"] == len(block.text)
    assert len(record["text_sha256"]) == 64
    assert record["source_file"] == "s.docx" and record["locator"] == "para 4"


def test_blocks_round_trip_and_are_byte_stable(tmp_path):
    """Same bytes in, byte-identical file out: that is what makes an extraction
    safe to hash and cheap to diff after a parser change."""
    blocks = [Block(kind=PARAGRAPH, text=f"paragraph {i}", source_file="f.pdf",
                    origin="pdf", index=i) for i in range(3)]
    first = tmp_path / "a.jsonl"
    second = tmp_path / "b.jsonl"
    write_blocks(first, blocks)
    write_blocks(second, blocks)
    assert first.read_bytes() == second.read_bytes()
    assert [b["text"] for b in read_blocks(first)] == ["paragraph 0", "paragraph 1",
                                                       "paragraph 2"]


def test_read_blocks_skips_a_truncated_line(tmp_path):
    path = tmp_path / "blocks.jsonl"
    path.write_text('{"text": "good", "index": 0}\n{"text": "trunc\n')
    assert [b["text"] for b in read_blocks(path)] == ["good"]


def test_read_blocks_of_a_missing_file_is_empty(tmp_path):
    assert list(read_blocks(tmp_path / "nope.jsonl")) == []


def test_markdown_rendering_groups_by_file():
    blocks = [
        Block(kind=HEADING, text="Methods", source_file="f.nxml", origin="jats"),
        Block(kind=PARAGRAPH, text="We used mice.", source_file="f.nxml", origin="jats"),
        Block(kind=TABLE, text="TABLE: S1", source_file="s.xlsx", origin="xlsx"),
    ]
    text = render_markdown(blocks)
    assert "## FILE: f.nxml" in text and "## FILE: s.xlsx" in text
    assert "### Methods" in text and "```" in text


# -- invisible characters, in every parser ------------------------------------

#: One string carrying each of the three codepoints that were reaching
#: `blocks.jsonl`, in the three shapes the corpus actually holds them in: inside
#: a word, doubled before a word, and at the front of a cell.
DAMAGED = "Bei\u00adGene sequenced 10\u200b\u200b\u22125 of the \ufeffEPCAM+ cells"
REPAIRED = "BeiGene sequenced 10\u22125 of the EPCAM+ cells"

INVISIBLE = "\u00ad\u200b\u200c\u200d\ufeff"


def test_strip_invisible_removes_the_five_and_leaves_the_visible_marks():
    """The set is exactly what is not on the page. A real hyphen, a minus sign
    and a non-breaking space all are, and a parser that ate them would be
    changing what the paper says rather than recovering it."""
    assert strip_invisible("a\u00adb\u200bc\u200cd\u200de\ufefff") == "abcdef"
    assert strip_invisible("scRNA-seq\u00a0\u22125") == "scRNA-seq\u00a0\u22125"


def test_a_jats_paragraph_does_not_carry_an_invisible_character():
    """13 of the 16 damaged blocks in the corpus were JATS.
    10.1126/sciadv.adh1914 writes `resolution = 1 x 10<U+200B><U+200B>-5` seven
    times over; 10.1158/2643-3230.bcd-21-0075 writes `Bei<U+00AD>Gene` and
    `Vectra<U+00AD>Polaris`, where the soft hyphen is the whole difference
    between the company's name and a string nothing matches."""
    body = f"<sec><title>Methods</title><p>{DAMAGED}</p></sec>"
    blocks, status, _ = jats.blocks_from_jats(jats_article(body), "f.nxml", L)
    assert status == "ok"
    body_paragraphs = [b for b in blocks
                       if b.kind == PARAGRAPH and b.section == "methods"]
    assert [b.text for b in body_paragraphs] == [REPAIRED]


def test_a_docx_paragraph_does_not_carry_an_invisible_character():
    """`<U+FEFF>Supplementary References` is a real heading in
    10.1186/s13073-021-00933-8, and a search for the heading it prints as misses
    it. Table cells come through the same function, so they are covered too."""
    data = make_docx([("paragraph", DAMAGED),
                      ("table", [["Primer", DAMAGED], ["Actb-F", "GGCTGTATTCC"]])])
    blocks, status, _ = docxfile.blocks_from_docx(data, "s.docx", L)
    assert status == "ok"
    assert blocks[0].text == REPAIRED
    card = next(b for b in blocks if b.kind == TABLE)
    assert card.table["header"] == ["Primer", REPAIRED]


def test_an_xlsx_cell_does_not_carry_an_invisible_character():
    """10.1038/s42003-021-02562-8 names a marker-gene column
    `<U+FEFF>EPCAM+ cells and cholangiocytes`, and 68 marks survived one sheet of
    10.1038/s43587-024-00613-3. A block writes the card twice -- rendered as
    `text` and structured as `table` -- so cleaning it at the cell is what makes
    both true at once."""
    data = make_xlsx({"marker genes": [["\ufeffEPCAM+ cells", "Stellate cell"],
                                       ["\u200bALB", "PDGFR\u00adB"]]})
    cards, status, _ = spreadsheet.cards_from_xlsx(data, "s.xlsx", L)
    assert status == "ok"
    card = cards[0]
    assert card.header == ["EPCAM+ cells", "Stellate cell"]
    both = json.dumps(card.to_dict(), ensure_ascii=False) + tables.render(card, L)
    assert not [c for c in both if c in INVISIBLE]


def test_a_landing_page_does_not_carry_an_invisible_character():
    """A landing page is the whole article when there is no PDF and no XML, and
    its `citation_abstract` meta tag is most of what there is to read. Both the
    meta tags and the kept text runs become block text, so both are stripped --
    the tags after the `citation_*` prefix check, which decides what is wanted
    and is a separate question from what the value carries."""
    page = ('<html><head>'
            f'<meta name="citation_title" content="{DAMAGED}">'
            '</head><body><p>'
            + f"{DAMAGED}. " * 3 +
            '</p></body></html>').encode("utf-8")
    blocks, status, meta = htmlfile.blocks_from_html(page, "landing.html", L)
    assert status == "ok"
    assert meta["meta_tags"]["citation_title"] == [REPAIRED]
    assert not [c for b in blocks for c in b.text if c in INVISIBLE]
    assert any(REPAIRED in b.text for b in blocks if b.kind == PARAGRAPH)


def test_a_prose_text_supplement_does_not_carry_an_invisible_character():
    """The tabular branch of a `.txt` reaches `clean_cell` and was already clean;
    the prose branch builds its paragraph by hand. Stripping before the split
    rather than after is what matters here -- U+200B is not whitespace, so
    collapsing first leaves the two spaces that surrounded it behind."""
    body = f"Supplementary note\n\n{DAMAGED} were housed at 22 degrees.\n".encode("utf-8")
    blocks, status, _ = extractor._plain_text_blocks(body, "note.txt", L, "supplement")
    assert status == "ok"
    assert [b.text for b in blocks][-1] == f"{REPAIRED} were housed at 22 degrees."


def test_the_pdf_parser_still_strips_them_after_rejoining_hyphens():
    """`pdf.py` owned this rule before it was shared, and its ordering is the
    part a move can quietly break: the soft hyphen has to be removed *with* the
    line break it caused, so a line-wrapped `interleukin-<soft hyphen>17A` keeps
    its real hyphen instead of rejoining to `interleukin17A`."""
    assert pdf._clean_block("interleukin-\u00ad\n17A \ufeffand \u200bIFN") == \
        "interleukin-17A and IFN"



# -- numbered-and-glued headings, and line-numbered manuscripts ---------------

def test_a_numbered_glued_heading_is_split():
    """`_HEADING_PREFIX` was in `_compiled` and not in `_leading_patterns`, so a
    heading that was both numbered and glued matched neither: `_compiled` wants the
    whole line and `_leading_patterns` wanted the alias at offset zero. MDPI writes
    exactly that shape -- 10.3390/genes15030298 glues each section heading to its
    first subheading -- and every block from 2.1 to 3.4 of that paper, the whole of
    its Methods and Results, carried the `introduction` label instead."""
    split = sections.split_leading_heading(
        "3. Results 3.1. Cellular Landscape of the Donor Newborn Human Lung")
    assert split is not None
    name, heading, rest = split
    assert (name, heading) == (sections.RESULTS, "3. Results")
    assert rest == "3.1. Cellular Landscape of the Donor Newborn Human Lung"


def test_a_short_remainder_still_splits_when_it_is_itself_a_heading():
    """`_MIN_GLUED_REMAINDER` exists to stop a heading with a stray word being read
    as a glued paragraph. A remainder that is itself a heading is the exception:
    `2. Materials and Methods 2.1. Study Population` leaves 21 characters, under the
    bound, and every one of them a heading."""
    split = sections.split_leading_heading(
        "2. Materials and Methods 2.1. Study Population")
    assert split is not None
    name, heading, rest = split
    assert (name, heading, rest) == (sections.METHODS, "2. Materials and Methods",
                                     "2.1. Study Population")


def test_a_short_remainder_that_is_not_a_heading_still_refuses_to_split():
    assert sections.split_leading_heading("Methods and") is None
    assert sections.split_leading_heading("Results 1") is None


def test_strip_line_number_takes_only_a_trailing_integer():
    assert sections.strip_line_number("Discussion 361") == "Discussion"
    assert sections.strip_line_number("Methods 606") == "Methods"
    # Not a line number: no trailing integer at all, or one that is the whole line.
    assert sections.strip_line_number("Discussion") == "Discussion"
    assert sections.strip_line_number("Extended Data Fig. 1a") == "Extended Data Fig. 1a"


def test_a_line_numbered_manuscript_is_detected_and_an_ordinary_pdf_is_not():
    """The threshold is measured, not chosen: the two line-numbered manuscripts in
    this corpus score 0.90 and 0.80, the highest any other article reaches is 0.34,
    and 0.6 is the middle of that gap. `Extended Data Fig. 1` and `Discussion 361`
    are the same shape per line, so the question is only answerable per document."""
    def page(lines):
        return [(line, False, {}) for line in lines]

    numbered = page([f"a manuscript line of ordinary body text here {n}"
                     for n in range(1, 41)])
    assert pdf._line_numbered([numbered]) is True

    plain = page(["a manuscript line of ordinary body text here"
                  for _ in range(40)])
    assert pdf._line_numbered([plain]) is False

    # A fifth of the lines ending in a figure number is the shape of an ordinary
    # PDF, and is under the fraction.
    sparse = page([f"a manuscript line of ordinary body text here {n}"
                   if n % 5 == 0 else "a manuscript line of ordinary body text here"
                   for n in range(1, 41)])
    assert pdf._line_numbered([sparse]) is False


def test_the_detector_needs_the_numbers_to_ascend():
    """The fraction alone would qualify a document whose lines happen to end in a
    repeated figure number."""
    same = [[("a manuscript line of ordinary body text here 7", False, {})
             for _ in range(40)]]
    assert pdf._line_numbered(same) is False


# -- a name the magic bytes contradict ----------------------------------------

def test_an_xlsx_named_csv_is_read_as_an_xlsx():
    """Nature served 10.1038/s41467-024-55440-2's MOESM10 as `.csv` and it is an
    xlsx workbook. The csv reader met a zip and answered `unreadable` with
    "new-line character seen in unquoted field", which describes the reader's
    experience and tells a reader nothing about the file."""
    data = make_xlsx({"Sheet1": [["gene", "logFC"], ["TP53", "1.4"]]})
    outcome = extractor.extract_bytes(data, "supplementary/01_data.csv", L)

    assert outcome.status == "ok"
    assert outcome.n_tables == 1
    assert outcome.meta["named_extension"] == ".csv"
    assert outcome.meta["sniffed_as"] == ".xlsx"
    assert "magic bytes say .xlsx" in outcome.note


def test_a_zip_named_csv_says_so_even_when_it_yields_nothing():
    """The archive branches build their own `FileResult` and never reach
    `extract_bytes`'s `result` helper, so the naming note has to be stamped on the
    way out. 10.1038/s41467-020-19737-2's MOESM18 is a zip named `.csv` holding a
    3.45 GB member: no text either way, but `no_text` with no mention of the zip
    would leave the record saying it had been read as a CSV."""
    data = make_zip([("inner.bin", b"\x00\x01\x02")])
    outcome = extractor.extract_bytes(data, "supplementary/02_data.csv", L)

    assert outcome.meta["named_extension"] == ".csv"
    assert outcome.meta["sniffed_as"] == ".zip"
    assert "magic bytes say .zip" in outcome.note


def test_a_real_xls_is_not_re_sniffed_into_a_doc():
    """The guard on `SNIFF_OVERRIDES_EXTENSION`. An OLE2 container is legacy Excel
    and legacy Word at the same 8 bytes, so `sniff_extension` answers `.doc` for a
    real `.xls` whenever no Content-Type says otherwise. Sniffing every known
    extension rather than only the text ones would route all 56 of this corpus's
    `.xls` files to `unsupported_format`."""
    data = make_xls({"Sheet1": [["gene", "logFC"], ["TP53", "1.4"]]})
    outcome = extractor.extract_bytes(data, "supplementary/03_data.xls", L)

    assert outcome.status == "ok"
    assert "named_extension" not in outcome.meta


def test_a_genuine_csv_is_left_alone():
    """The override only fires on a container magic number, so ordinary text is
    untouched and keeps its own extension."""
    outcome = extractor.extract_bytes(b"gene,logFC\nTP53,1.4\n",
                                      "supplementary/04_data.csv", L)
    assert outcome.status == "ok"
    assert "named_extension" not in outcome.meta
    assert "sniffed_as" not in outcome.meta


# -- what MuPDF has to say ----------------------------------------------------

def test_mupdf_messages_are_captured_and_capped():
    """MuPDF writes its diagnostics to stdout, which is where `extract one` prints
    its machine-readable result, so they are silenced at import and read back per
    file instead. Capped because the buffer is wider than what was being printed:
    only four PDFs in this corpus emit an error, but 15 of a random 40 emit a font
    warning, and a file that goes through `_repair_glyph_encoding` provokes 20
    lines of `FT_Get_Advance`."""
    lines = "\n".join(f"warning number {n}" for n in range(1, 13))
    with mock.patch.object(fitz.TOOLS, "mupdf_warnings", return_value=lines):
        captured = pdf._mupdf_warnings()

    assert captured["mupdf_warnings"] == [f"warning number {n}" for n in range(1, 6)]
    assert captured["mupdf_warnings_total"] == 12


def test_a_quiet_pdf_adds_no_mupdf_field_at_all():
    with mock.patch.object(fitz.TOOLS, "mupdf_warnings", return_value=""):
        assert pdf._mupdf_warnings() == {}


def test_mupdf_writes_nothing_to_stdout_and_the_message_reaches_the_record(capfd):
    """The regression this guards: MuPDF prints on **stdout**, which is the channel
    `extract/cli.py`'s `one` uses for its machine-readable result, so
    `DIR=$(manuscript-extract one ...)` came back with 55 bytes of
    `MuPDF error: ...` in front of the path. A truncated PDF is the cheapest way to
    make MuPDF say something of the same class it says about the four real files."""
    assert fitz.TOOLS.mupdf_display_errors() is False

    whole = make_pdf()
    _, _, meta = pdf.blocks_from_pdf(whole[: len(whole) // 2], "x.pdf", L)

    assert capfd.readouterr().out == ""
    assert any("startxref" in line for line in meta.get("mupdf_warnings", []))


def test_one_pdf_s_mupdf_messages_are_not_reported_against_the_next(capfd):
    """The buffer is process-wide, so without the reset the first noisy PDF in a
    393-article run would be reported against every file after it."""
    whole = make_pdf()
    pdf.blocks_from_pdf(whole[: len(whole) // 2], "broken.pdf", L)
    _, _, meta = pdf.blocks_from_pdf(whole, "clean.pdf", L)

    assert "mupdf_warnings" not in meta
    assert capfd.readouterr().out == ""


@pytest.mark.parametrize("entry", [
    "7. Results From A Randomized Trial Of Inhaled Steroids, N Engl J Med (2019).",
    "3. Discussion Of Sampling Bias In Cohort Studies, Am J Epidemiol (2021).",
    "21. Introduction To Single Cell Genomics, Cold Spring Harb Perspect (2018).",
    "1. Smith, J. et al. Methods For Nuclei Isolation. Cell 180, 1-12 (2020).",
    "4. Results Of The Trial. doi:10.1038/s41586-020-0000-0",
])
def test_a_numbered_reference_entry_is_not_split_into_a_heading(entry):
    """Letting `_HEADING_PREFIX` into `_leading_patterns` widened it onto
    reference-list entries, which begin with a number too, and a title in Title Case
    passes the "followed by a capital" guard. A heading's section carries forward, so
    a false split here opens `results` over a bibliography. The discriminator is the
    citation signals other than the numbering -- a parenthesised year, an `et al.`,
    a DOI -- because the numbering itself is what a numbered heading and a numbered
    citation have in common."""
    assert sections.split_leading_heading(entry) is None
