"""Spreadsheets -> table cards. The largest supplement category in the corpus.

172 of the supplementary files here are `.xlsx`, one is a legacy `.xls`, and 18
are `.csv`. Two things this module refuses to do, both learned from the corpus:

- It never calls `calculate_dimension()`. `44161_2025_612_MOESM5_ESM.xlsx` has
  worksheets with no declared dimensions, and openpyxl raises
  `ValueError: Worksheet is unsized` for those; the reader has to reset the
  dimensions and count rows itself.
- It never reads a whole sheet just to find out how big it is. Scanning stops at
  `limits.max_scan_rows` and the card says the profile is partial.
"""

import csv
import hashlib
import io
import warnings
from typing import Any, List, Sequence, Tuple

from . import ooxml, tables
from .limits import Limits

OK = "ok"
NO_TEXT = "no_text"
UNREADABLE = "unreadable"
UNSUPPORTED = "unsupported_format"

_DELIMITERS = ",\t;|"


def _scan_rows(iterator, limits: Limits) -> Tuple[List[Sequence[Any]], bool]:
    """Pull up to the cap, plus one row to learn whether more existed."""
    rows: List[Sequence[Any]] = []
    truncated = False
    for row in iterator:
        if len(rows) >= limits.max_scan_rows:
            truncated = True
            break
        rows.append(tuple(row))
    return rows, truncated


def _silence_openpyxl() -> None:
    """Drop openpyxl's "extension is not supported" warnings for this call.

    Publisher workbooks carry conditional-formatting, data-validation and chart
    extensions that openpyxl warns about and then discards. They concern features
    this stage never reads. The filter has to be installed around the *row
    iteration* as well as the load, because a read-only workbook parses its
    sheets lazily and the warnings fire while scanning.
    """
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


def _load(data: bytes):
    """Open a workbook, with the strict-OOXML retry. Raises on unreadable bytes."""
    import openpyxl

    workbook = openpyxl.load_workbook(
        io.BytesIO(data), read_only=True, data_only=True, keep_links=False)
    if workbook.worksheets:
        return workbook, False
    # Zero worksheets is how openpyxl reports a strict-conformance workbook.
    # Relax the namespaces and try once more before believing it is empty.
    workbook.close()
    relaxed = ooxml.relax_strict(data)
    if relaxed is None:
        return openpyxl.load_workbook(
            io.BytesIO(data), read_only=True, data_only=True, keep_links=False), False
    return openpyxl.load_workbook(
        io.BytesIO(relaxed), read_only=True, data_only=True, keep_links=False), True


def _reset_dimensions(sheet) -> None:
    """Make an unsized worksheet yield rows. Silent when the reader cannot.

    `44161_2025_612_MOESM5_ESM.xlsx` has worksheets with no declared dimensions, and
    a read-only openpyxl sheet yields nothing at all for those until this is called.
    The three exception types are what different openpyxl versions raise for a sheet
    that has no dimensions to reset, none of which is a problem with the file.
    """
    try:
        sheet.reset_dimensions()
    except (AttributeError, TypeError, ValueError):
        pass


def _apply_table_cap(cards: List[tables.TableCard], limits: Limits, meta: dict) -> bool:
    """Trim to `max_tables_per_file`, recording it. True when the caller should stop.

    Shared by the .xlsx and .xls readers; the .xls copy of this was uncovered, so a
    change to how the cap is reported would have reached one format and not the
    other. `setdefault` on `reason` is deliberate -- an earlier, more specific
    reason (a sheet that failed to scan) outranks this one.
    """
    if len(cards) < limits.max_tables_per_file:
        return False
    del cards[limits.max_tables_per_file:]
    meta["tables_capped"] = True
    meta.setdefault(
        "reason",
        f"stopped at the {limits.max_tables_per_file}-table cap; later "
        f"sheets in this workbook were not profiled")
    return True


def _forced(overrides, source_file: str, locator: str) -> dict:
    """A human's answer about this card's header row, as build_card keywords.

    The translation itself is `review.Overrides.header_kwargs`; this only tolerates
    an article with no review file at all.
    """
    if overrides is None:
        return {}
    return overrides.header_kwargs(source_file, locator)


def _cards_from_sheet(rows: List[Sequence[Any]], source_file: str, title: str,
                      limits: Limits, total, truncated: bool, meta: dict,
                      sha256: str = "", overrides=None) -> List[tables.TableCard]:
    """One sheet's cards: normally one, but one per panel when it holds several.

    A sheet of stacked panels used to become a single card that pooled them --
    six experiments' control and treatment columns under the first panel's
    header. `tables.split_blocks` says when that is what is happening.
    """
    parts = tables.split_blocks(rows, limits)
    if not parts:
        locator = f"sheet {title!r}"
        card = tables.build_card(
            rows, source_file=source_file, locator=locator, limits=limits,
            title=title, n_rows_total=total, truncated=truncated,
            data_ref={"file": source_file, "sheet": title, "sha256": sha256},
            **_forced(overrides, source_file, locator),
        )
        return [card] if card is not None else []

    kept = parts[: limits.max_tables_per_sheet]
    if len(parts) > len(kept):
        meta["tables_skipped"] = meta.get("tables_skipped", 0) + len(parts) - len(kept)
    built: List[tables.TableCard] = []
    for start, end in kept:
        locator = f"sheet {title!r} rows {start + 1}-{end}"
        card = tables.build_card(
            rows[start:end], source_file=source_file,
            locator=locator, limits=limits,
            title=title, n_rows_total=None,
            # Only the part that runs to the end of the scanned window is the one
            # the cap actually cut; the panels above it were read in full.
            truncated=truncated and end == len(rows),
            data_ref={"file": source_file, "sheet": title, "sha256": sha256,
                      "row_start": start + 1, "row_end": end},
            row_offset=start,
            **_forced(overrides, source_file, locator),
        )
        if card is None:
            continue
        card.notes.insert(0, f"sheet {title!r} holds {len(parts)} blank-row-separated "
                             f"tables; this card is rows {start + 1}-{end}")
        built.append(card)
    return built


def cards_from_xlsx(
    data: bytes, source_file: str, limits: Limits, overrides=None
) -> Tuple[List[tables.TableCard], str, dict]:
    try:
        import openpyxl  # noqa: F401
    except ImportError:  # pragma: no cover - openpyxl is a hard requirement
        return [], UNSUPPORTED, {"reason": "openpyxl is not installed"}

    with warnings.catch_warnings():
        _silence_openpyxl()
        return _cards_from_xlsx(data, source_file, limits, overrides)


def _cards_from_xlsx(
    data: bytes, source_file: str, limits: Limits, overrides=None
) -> Tuple[List[tables.TableCard], str, dict]:
    try:
        workbook, relaxed = _load(data)
    except Exception as e:
        return [], UNREADABLE, {"reason": f"{type(e).__name__}: {e}"}

    sha256 = hashlib.sha256(data).hexdigest()

    cards: List[tables.TableCard] = []
    meta: dict = {"sheets": 0, "sheets_skipped": 0}
    if relaxed:
        meta["strict_ooxml"] = True
    try:
        worksheets = list(workbook.worksheets)
        meta["sheets"] = len(worksheets)
        if not worksheets:
            return [], UNREADABLE, {
                **meta,
                "reason": "the workbook declares no worksheets; if it is a strict "
                          "ISO-29500 file the namespace relaxation did not help"}
        if len(worksheets) > limits.max_sheets:
            meta["sheets_skipped"] = len(worksheets) - limits.max_sheets
            worksheets = worksheets[: limits.max_sheets]

        for sheet in worksheets:
            # Read the declared row count first: `reset_dimensions` clears it, and
            # in read-only mode nothing recomputes it, so asking afterwards gives
            # None even for a sheet that said how big it was.
            declared = sheet.max_row
            total = declared if isinstance(declared, int) and declared > 0 else None
            _reset_dimensions(sheet)
            try:
                rows, truncated = _scan_rows(sheet.iter_rows(values_only=True), limits)
            except Exception as e:
                meta.setdefault("errors", []).append(
                    f"sheet {sheet.title!r}: {type(e).__name__}: {e}")
                continue

            cards.extend(_cards_from_sheet(rows, source_file, sheet.title, limits,
                                           total, truncated, meta, sha256,
                                           overrides))
            if _apply_table_cap(cards, limits, meta):
                break
    finally:
        workbook.close()

    if not cards:
        return [], NO_TEXT, meta
    return cards, OK, meta


def cards_from_xls(
    data: bytes, source_file: str, limits: Limits, overrides=None
) -> Tuple[List[tables.TableCard], str, dict]:
    """Legacy binary `.xls`. One file in the 63-paper development corpus needs it.

    `xlrd` 2.x reads only this format, which is exactly the reason to keep it
    optional: without it the single file is reported as unsupported rather than
    dragging in a dependency for 1 of 191 spreadsheets.

    That file is not in a fetched corpus here, so this path is tested against a
    stub standing in for `xlrd` rather than against real bytes.
    """
    try:
        import xlrd
    except ImportError:
        return [], UNSUPPORTED, {
            "reason": "legacy .xls needs the optional xlrd package (pip install xlrd)"}

    try:
        book = xlrd.open_workbook(file_contents=data)
    except Exception as e:
        return [], UNREADABLE, {"reason": f"{type(e).__name__}: {e}"}

    cards: List[tables.TableCard] = []
    meta: dict = {"sheets": book.nsheets}
    sha256 = hashlib.sha256(data).hexdigest()
    for sheet in book.sheets()[: limits.max_sheets]:
        limit = min(sheet.nrows, limits.max_scan_rows)
        rows = [tuple(sheet.row_values(index)) for index in range(limit)]
        cards.extend(_cards_from_sheet(rows, source_file, sheet.name, limits,
                                       sheet.nrows, sheet.nrows > limit, meta,
                                       sha256, overrides))
        if _apply_table_cap(cards, limits, meta):
            break
    if not cards:
        return [], NO_TEXT, meta
    return cards, OK, meta


def _source_rows(data: bytes, data_ref: dict, extension: str):
    """Yield `(row_number, cells)` from the source a `data_ref` points at.

    Row numbers are 1-based, matching the numbers `build_card` writes.
    """
    if extension in {".xls"}:
        import xlrd
        book = xlrd.open_workbook(file_contents=data)
        sheet = book.sheet_by_name(data_ref["sheet"])
        for index in range(sheet.nrows):
            yield index + 1, list(sheet.row_values(index))
        return
    if data_ref.get("sheet"):
        with warnings.catch_warnings():
            _silence_openpyxl()
            workbook, _ = _load(data)
            try:
                sheet = workbook[data_ref["sheet"]]
                _reset_dimensions(sheet)
                for number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    yield number, list(row)
            finally:
                workbook.close()
        return
    text = data.decode("utf-8-sig", errors="replace")
    delimiter = data_ref.get("delimiter") or _sniff_delimiter(text[:8192])
    for number, row in enumerate(csv.reader(io.StringIO(text), delimiter=delimiter),
                                 start=1):
        yield number, list(row)


def read_rows(data: bytes, data_ref: dict, extension: str,
              limit: int = 20) -> Tuple[List[str], List[Tuple[int, List[str]]]]:
    """Re-read the header and data rows a card describes. `([], [])` when it cannot.

    This is what makes `data_ref` a contract rather than a comment: the module
    docstring has always said a card "records `data_ref` -- file, sheet, header
    row -- so code that wants the real values re-reads the original file at the
    exact offset", and until `manuscript-extract table` existed nothing did.
    """
    first = data_ref.get("first_data_row")
    if not first:
        return [], []
    last = data_ref.get("last_data_row") or first
    header_at = data_ref.get("header_row")
    stop = min(last, first + limit - 1)

    header: List[str] = []
    found: List[Tuple[int, List[str]]] = []
    for number, cells in _source_rows(data, data_ref, extension):
        if number == header_at:
            header = [tables.clean_cell(c) or "" for c in cells]
        elif first <= number <= stop:
            found.append((number, [tables.clean_cell(c) or "" for c in cells]))
        if number >= stop and (header or header_at is None):
            break
    return header, found


def _sniff_delimiter(sample: str) -> str:
    try:
        return csv.Sniffer().sniff(sample, delimiters=_DELIMITERS).delimiter
    except csv.Error:
        # Fall back to whichever candidate appears most on the first line.
        first = sample.splitlines()[0] if sample.splitlines() else ""
        counts = {d: first.count(d) for d in _DELIMITERS}
        best = max(counts, key=counts.get)
        return best if counts[best] else ","


def cards_from_csv(
    data: bytes, source_file: str, limits: Limits, overrides=None
) -> Tuple[List[tables.TableCard], str, dict]:
    text = data.decode("utf-8-sig", errors="replace")
    if not text.strip():
        return [], NO_TEXT, {}
    delimiter = _sniff_delimiter(text[:8192])
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    try:
        rows, truncated = _scan_rows(reader, limits)
        # Count the file exactly. `n_rows_total=None` made
        # 10.1126/science.aat5031's 40,269-line data_s1.csv read
        # `Shape: 4998 data row(s) x 7 column(s)` with no total -- indistinguishable
        # on that line from the 60-row file beside it. The whole text is already
        # decoded above so this costs no I/O, and counting parsed rows rather than
        # newlines keeps a quoted embedded newline from inflating the figure. The
        # "never read a whole sheet to size it" rule is about xlsx, where sizing
        # would mean a full parse of bytes nobody has touched yet.
        total = sum(1 for row in csv.reader(io.StringIO(text), delimiter=delimiter)
                    if any(str(cell).strip() for cell in row))
    except csv.Error as e:
        return [], UNREADABLE, {"reason": f"{type(e).__name__}: {e}"}

    card = tables.build_card(
        rows,
        source_file=source_file,
        locator="rows",
        limits=limits,
        n_rows_total=total,
        truncated=truncated,
        data_ref={"file": source_file, "delimiter": delimiter,
                  "sha256": hashlib.sha256(data).hexdigest()},
        **_forced(overrides, source_file, "rows"),
    )
    meta = {"delimiter": delimiter}
    if card is None:
        return [], NO_TEXT, meta
    return [card], OK, meta
